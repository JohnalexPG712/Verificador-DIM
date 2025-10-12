import os
import re
import streamlit as st
import pandas as pd
import pdfplumber
import glob
from datetime import datetime
import tempfile
import io
import numpy as np
from collections import OrderedDict
from openpyxl import load_workbook
import warnings
import unicodedata

# Configuración de página
st.set_page_config(
    page_title="Sistema de Validación de Importaciones", 
    page_icon="📊", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ocultar advertencias
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# =============================================================================
# CLASE PARA CORRECCIÓN DE NOMBRES (Del primer código)
# =============================================================================

class CorrectorNombres:
    """Clase simplificada para corregir nombres basada en número de letras"""
    
    def normalizar_texto(self, texto):
        """Normaliza texto removiendo todo excepto letras"""
        if not texto or texto == "NO ENCONTRADO":
            return ""
        
        # Convertir a mayúsculas y remover tildes
        texto = str(texto).upper()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto) 
                       if unicodedata.category(c) != 'Mn')
        
        # Remover TODO excepto letras (espacios, puntos, números, caracteres especiales)
        texto = re.sub(r'[^A-Z]', '', texto)
        
        return texto
    
    def comparar_por_letras(self, nombre_pdf, nombre_excel):
        """Compara dos nombres por número de letras (sin importar espacios, puntos, etc.)"""
        if not nombre_pdf or not nombre_excel:
            return False
        
        # Normalizar ambos textos (solo letras)
        pdf_normalizado = self.normalizar_texto(nombre_pdf)
        excel_normalizado = self.normalizar_texto(nombre_excel)
        
        # Comparar si tienen el mismo número de letras
        return len(pdf_normalizado) == len(excel_normalizado)
    
    def corregir_nombre(self, nombre_pdf, nombre_excel):
        """Corrige el nombre del PDF usando el Excel como referencia si coinciden en número de letras"""
        if not nombre_excel or nombre_excel == "NO ENCONTRADO":
            return nombre_pdf
        
        # Comparar por número de letras
        if self.comparar_por_letras(nombre_pdf, nombre_excel):
            return nombre_excel
        else:
            return nombre_pdf

# =============================================================================
# CLASE PARA VALIDACIÓN DIM vs ANEXOS FMM (Del primer código)
# =============================================================================

class ValidadorDeclaracionImportacionCompleto:
    def __init__(self):
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        
        # Inicializar el corrector de nombres simplificado
        self.corrector_nombres = CorrectorNombres()
        
        self.CAMPOS_DI = {
            "5.": "5. Número de Identificación Tributaria (NIT)",
            "11.": "11. Apellidos y Nombres / Razón Social Importador",
            "42.": "42. No. Manifiesto de Carga",
            "43.": "43. Fecha Manifiesto de Carga",
            "44.": "44. No. Documento de Transporte",
            "45.": "45. Fecha Documento de Transporte",
            "51.": "51. No. Factura Comercial",
            "52.": "52. Fecha Factura Comercial",
            "132.": "132. No. Aceptación Declaración",
            "133.": "133. Fecha Aceptación",
            "134.": "134. Levante No.",
            "135.": "135. Fecha Levante"
        }

        self.MAPEOS_VALIDACION = {
            "5. Número de Identificación Tributaria (NIT)": {
                "codigo_formulario": "PROVEEDOR",
                "descripcion_esperada": "INFORMACION_PROVEEDOR",
                "tipo": "documento",
                "cambia_por_declaracion": False
            },
            "11. Apellidos y Nombres / Razón Social Importador": {
                "codigo_formulario": "PROVEEDOR", 
                "descripcion_esperada": "INFORMACION_PROVEEDOR",
                "tipo": "documento",
                "cambia_por_declaracion": False
            },
            "42. No. Manifiesto de Carga": {
                "codigo_formulario": 93,
                "descripcion_esperada": "FORMULARIO DE SALIDA ZONA FRANCA",
                "tipo": "documento",
                "cambia_por_declaracion": False
            },
            "43. Fecha Manifiesto de Carga": {
                "codigo_formulario": 93,
                "descripcion_esperada": "FORMULARIO DE SALIDA ZONA FRANCA",
                "tipo": "fecha",
                "cambia_por_declaracion": False
            },
            "44. No. Documento de Transporte": {
                "codigo_formulario": 17,
                "descripcion_esperada": "DOCUMENTO OF TRANSPORTE",
                "tipo": "documento",
                "cambia_por_declaracion": False
            },
            "45. Fecha Documento de Transporte": {
                "codigo_formulario": 17,
                "descripcion_esperada": "DOCUMENTO OF TRANSPORTE",
                "tipo": "fecha",
                "cambia_por_declaracion": False
            },
            "51. No. Factura Comercial": {
                "codigo_formulario": 6,
                "descripcion_esperada": "FACTURA COMERCIAL",
                "tipo": "documento",
                "cambia_por_declaracion": True
            },
            "52. Fecha Factura Comercial": {
                "codigo_formulario": 6,
                "descripcion_esperada": "FACTURA COMERCIAL",
                "tipo": "fecha",
                "cambia_por_declaracion": True
            },
            "132. No. Aceptación Declaración": {
                "codigo_formulario": 9,
                "descripcion_esperada": "DECLARACION DE IMPORTACION",
                "tipo": "documento",
                "cambia_por_declaracion": True
            },
            "133. Fecha Aceptación": {
                "codigo_formulario": 9,
                "descripcion_esperada": "DECLARACION DE IMPORTACION",
                "tipo": "fecha",
                "cambia_por_declaracion": True
            },
            "134. Levante No.": {
                "codigo_formulario": 47,
                "descripcion_esperada": "AUTORIZACION DE LEVANTE",
                "tipo": "documento",
                "cambia_por_declaracion": True
            },
            "135. Fecha Levante": {
                "codigo_formulario": 47,
                "descripcion_esperada": "AUTORIZACION DE LEVANTE",
                "tipo": "fecha",
                "cambia_por_declaracion": True
            }
        }

        self.patrones = {
            "5. Número de Identificación Tributaria (NIT)": [
                r"5\s*\.?\s*N[uú]mero\s*de\s*Identificaci[oó]n\s*Tributaria\s*\(NIT\).*?([0-9]{6,12})",
                r"5\.\s*Número de Identificación Tributaria \(NIT\)[\s\S]*?(\d{6,12})"
            ],
            "11. Apellidos y Nombres / Razón Social Importador": [
                r"11\s*\.?\s*Apellidos\s*y\s*nombres\s*o\s*Raz[oó]n\s*Social\s*\n?\s*\d{6,12}\s*\d?\s*([A-ZÁÉÍÓÚÑ0-9\s\.\-]+?)(?=\s*13\s*\.)",
                r"11\.\s*Apellidos y nombres o Razón Social[\s\S]*?\n\s*(\d{6,12}\s*\d?\s*[A-ZÁÉÍÓÚÑ0-9\s\.\-]+)"
            ],
            "42. No. Manifiesto de Carga": [
                r"42\s*\.?\s*Manifiesto\s*de\s*carga[\s\S]*?No\.?\s*([A-Z0-9]+)"
            ],
            "43. Fecha Manifiesto de Carga": [
                r"43\s*\.?\s*Año\s*[-\s]*Mes\s*[-\s]*Día.*?(\d{4}\s*[-]\s*\d{2}\s*[-]\s*\d{2})"
            ],
            "44. No. Documento de Transporte": [
                r"44\s*\.?\s*Documento\s*de\s*transporte[\s\S]*?(?:No\.?\s*)?((?:[A-Z]+[0-9]+(?:\-[0-9]+)?)|(?:[0-9]{10,11}))(?=\s|[0-9]{4}|$)"
            ],
            "45. Fecha Documento de Transporte": [
                r"45\s*\.?\s*Año.*?Día[\s\S]*?[0-9]{4}\s*-\s*[0-9]{2}\s*-\s*[0-9]{2}[\s\S]*?([0-9]{4}\s*-\s*[0-9]{2}\s*-\s*[0-9]{2})"
            ],
            "51. No. Factura Comercial": [
                r"51\s*\.?\s*No\.?\s*de\s*factura[\s\S]*?\n\s*([A-Z0-9\-]+(?:\s*/\s*[A-Z0-9]+)?)"
            ],
            "52. Fecha Factura Comercial": [
                r"52\s*\.\s*?Año\s*-\s*Mes\s*-\s*Día.*?\n(?:.*?[^\d\w-])?(\d{4}\s*-\s*\d{2}\s*-\s*\d{2})"
            ],
            "132. No. Aceptación Declaración": [
                r"132\s*\.?\s*No\.?\s*Aceptaci[oó]n\s*declaraci[oó]n[\s\S]*?(\d{12,18})"
            ],
            "133. Fecha Aceptación": [
                r"133\s*\.?\s*Fec*h?a:?\s*(\d{4}\s*[\-\s]*\d{2}\s*[\-\s]*\d{2}|\d{8})\b"
            ],
            "134. Levante No.": [
                r"134\s*\.?\s*Levante\s*No\.?\s*(\d+)"
            ],
            "135. Fecha Levante": [
                r"135\s*\.?\s*Fec*h?a[\s\S]*?([\d]{4}\s*-\s*[\d]{2}\s*-\s*[\d]{2})"
            ]
        }

        self.nit_proveedor = None
        self.nombre_proveedor = None
        self.facturas_emparejadas = {}
        self._cache_nombres = {}  # Cache para optimizar comparaciones de nombres

    def buscar_archivo_formulario(self, carpeta):
        """Busca específicamente el archivo del formulario FMM"""
        st.info(f"🔍 Buscando formulario FMM...")
        
        patrones_formulario = [
            "*Rpt_Impresion_Formulario*",
            "*FORMULARIO*", 
            "*FMM*",
            "*.xlsx"
        ]
        
        for patron in patrones_formulario:
            archivos = glob.glob(os.path.join(carpeta, patron))
            for archivo in archivos:
                nombre_archivo = os.path.basename(archivo)
                if "Cruce" not in nombre_archivo and "validacion" not in nombre_archivo.lower():
                    st.success(f"📁 Formulario encontrado: {nombre_archivo}")
                    return archivo
        
        st.error("❌ No se encontró archivo del formulario FMM")
        return None

    def extraer_proveedor_formulario(self, archivo_excel):
        """Extrae NIT y nombre del proveedor del formulario FMM"""
        try:
            st.info(f"👤 Extrayendo información del proveedor...")
            
            wb = load_workbook(archivo_excel, data_only=True)
            sheet = wb.active
            
            proveedor_encontrado = False
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and 'Proveedor/Cliente:' in str(cell.value):
                        texto = str(cell.value)
                        st.info(f"📋 Información encontrada: {texto}")
                        
                        texto_limpio = texto.replace('Proveedor/Cliente:', '').strip()
                        
                        for i, char in enumerate(texto_limpio):
                            if char == ' ' and texto_limpio[:i].replace(' ', '').isdigit():
                                self.nit_proveedor = texto_limpio[:i].replace(' ', '').replace('-', '').replace('.', '')
                                self.nombre_proveedor = texto_limpio[i:].strip(' -')
                                break
                        
                        if not self.nit_proveedor and ' - ' in texto_limpio:
                            partes = texto_limpio.split(' - ', 1)
                            self.nit_proveedor = partes[0].replace(' ', '').replace('-', '').replace('.', '')
                            self.nombre_proveedor = partes[1]
                        
                        if self.nit_proveedor and self.nombre_proveedor:
                            proveedor_encontrado = True
                            st.success(f"✅ PROVEEDOR VÁLIDO:")
                            st.success(f"   🆔 NIT: {self.nit_proveedor}")
                            st.success(f"   📛 Nombre: {self.nombre_proveedor}")
                        else:
                            st.error(f"❌ ERROR: Formato de proveedor incorrecto")
                        
                        wb.close()
                        return proveedor_encontrado
            
            if not proveedor_encontrado:
                st.error("❌ ERROR: No se pudo extraer información del proveedor")
            
            wb.close()
            return proveedor_encontrado
            
        except Exception as e:
            st.error(f"❌ ERROR al extraer proveedor: {e}")
            return False

    def extraer_anexos_formulario_robusto(self, archivo_excel):
        """Extrae anexos del formulario de manera robusta con validación de duplicados"""
        try:
            st.info(f"📖 Extrayendo anexos del formulario...")
            
            wb = load_workbook(archivo_excel, data_only=True)
            sheet = wb.active
            
            inicio_anexos = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and 'DETALLE DE LOS ANEXOS' in str(cell_value):
                        inicio_anexos = row
                        break
                if inicio_anexos:
                    break
            
            if inicio_anexos is None:
                st.error("❌ ERROR: No se encontró la sección 'DETALLE DE LOS ANEXOS'")
                wb.close()
                return pd.DataFrame()
            
            encabezados = {}
            fila_encabezados = inicio_anexos + 1
            
            for col in range(1, sheet.max_column + 1):
                valor = sheet.cell(row=fila_encabezados, column=col).value
                if valor:
                    valor_str = str(valor).strip().upper()
                    if 'CÓDIGO' in valor_str or 'CODIGO' in valor_str:
                        encabezados['codigo'] = col
                    elif 'DESCRIPCIÓN' in valor_str or 'DESCRIPCION' in valor_str:
                        encabezados['descripcion'] = col
                    elif 'DOCUMENTO' in valor_str:
                        encabezados['documento'] = col
                    elif 'FECHA' in valor_str:
                        encabezados['fecha'] = col
            
            if not encabezados:
                encabezados = {
                    'codigo': 1,
                    'descripcion': 5,
                    'documento': 19,
                    'fecha': 34
                }
            
            datos_anexos = []
            fila_actual = fila_encabezados + 1
            
            for i in range(200):
                try:
                    if 'codigo' in encabezados:
                        codigo = sheet.cell(row=fila_actual, column=encabezados['codigo']).value
                    else:
                        codigo = sheet.cell(row=fila_actual, column=1).value
                    
                    if codigo is None or codigo == '':
                        filas_vacias = 0
                        for j in range(5):
                            codigo_check = sheet.cell(row=fila_actual + j, column=encabezados.get('codigo', 1)).value
                            if codigo_check is None or codigo_check == '':
                                filas_vacias += 1
                        if filas_vacias >= 3:
                            break
                        fila_actual += 1
                        continue
                    
                    try:
                        codigo_str = str(codigo).strip().split('.')[0]
                        if codigo_str not in ['6', '9', '17', '47', '93']:
                            fila_actual += 1
                            continue
                    except:
                        fila_actual += 1
                        continue
                    
                    if 'descripcion' in encabezados:
                        descripcion = sheet.cell(row=fila_actual, column=encabezados['descripcion']).value
                    else:
                        descripcion = sheet.cell(row=fila_actual, column=5).value
                    
                    if 'documento' in encabezados:
                        documento = sheet.cell(row=fila_actual, column=encabezados['documento']).value
                    else:
                        documento = sheet.cell(row=fila_actual, column=19).value
                    
                    if 'fecha' in encabezados:
                        fecha = sheet.cell(row=fila_actual, column=encabezados['fecha']).value
                    else:
                        fecha = sheet.cell(row=fila_actual, column=34).value
                    
                    fecha_normalizada = self.normalizar_fecha_dd_mm_aaaa(fecha, es_fecha=True)
                    
                    datos_anexos.append({
                        'Codigo': int(float(codigo_str)),
                        'Descripcion': descripcion,
                        'Documento': documento,
                        'Fecha': fecha_normalizada,
                        'Fila_Excel': fila_actual,
                        'Usado': False
                    })
                    
                    fila_actual += 1
                    
                except Exception:
                    fila_actual += 1
                    continue
            
            wb.close()
            
            df_resultado = pd.DataFrame(datos_anexos)
            
            if not df_resultado.empty:
                st.success(f"✅ {len(df_resultado)} anexos encontrados")
                
                # SIEMPRE mostrar resumen por código
                resumen = df_resultado.groupby('Codigo').agg({
                    'Descripcion': 'first',
                    'Documento': 'count'
                }).reset_index()
                
                st.info("📊 Resumen por código:")
                for _, row in resumen.iterrows():
                    st.info(f"   • Código {row['Codigo']}: {row['Documento']} - {row['Descripcion']}")
                
                # ✅ VALIDACIÓN: Solo mostrar errores si existen
                count_di = len(df_resultado[df_resultado['Codigo'] == 9])
                count_levante = len(df_resultado[df_resultado['Codigo'] == 47])
                
                # Verificar duplicados
                di_duplicados = df_resultado[
                    (df_resultado['Codigo'] == 9) & 
                    (df_resultado.duplicated('Documento', keep=False))
                ]
                levante_duplicados = df_resultado[
                    (df_resultado['Codigo'] == 47) & 
                    (df_resultado.duplicated('Documento', keep=False))
                ]
                
                has_errors = False
                mensajes_error = []
                
                if not di_duplicados.empty:
                    has_errors = True
                    documentos_duplicados = di_duplicados['Documento'].unique()
                    mensajes_error.append(f"❌ {len(documentos_duplicados)} DI duplicadas: {', '.join(documentos_duplicados)}")
                
                if not levante_duplicados.empty:
                    has_errors = True
                    documentos_duplicados = levante_duplicados['Documento'].unique()
                    mensajes_error.append(f"❌ {len(documentos_duplicados)} Levantes duplicados: {', '.join(documentos_duplicados)}")
                
                if count_di != count_levante:
                    has_errors = True
                    mensajes_error.append(f"❌ Desbalance: {count_di} DI vs {count_levante} Levantes")
                
                # Mostrar errores si existen
                if has_errors:
                    st.error("\n🔍 VALIDACIÓN DE INTEGRIDAD:")
                    for mensaje in mensajes_error:
                        st.error(f"   {mensaje}")
                else:
                    st.success(f"✅ Balance correcto: {count_di} DI = {count_levante} Levantes")
                    
            else:
                st.error("❌ No se encontraron anexos")
            
            return df_resultado
            
        except Exception as e:
            st.error(f"❌ Error al extraer anexos: {e}")
            return pd.DataFrame()

    def extraer_todas_declaraciones_pdf(self, pdf_path):
        """Extrae TODAS las declaraciones del PDF"""
        st.info(f"\n📄 Procesando PDF: {os.path.basename(pdf_path)}")
        
        texto_completo = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text(x_tolerance=3, y_tolerance=3)
                    if texto:
                        texto_completo += texto + "\n\n"
        except Exception as e:
            st.error(f"❌ Error al abrir el PDF: {e}")
            return []
        
        patron_declaraciones = r"4\s*\.?\s*N[uú]mero\s*de\s*formulario[\s\S]*?(\d{12,18})"
        matches = list(re.finditer(patron_declaraciones, texto_completo, re.IGNORECASE))
        
        st.info(f"📋 Declaraciones encontradas: {len(matches)}")
        
        declaraciones = []
        
        for i, match in enumerate(matches):
            numero_formulario = match.group(1)
            
            start_pos = match.start()
            if i < len(matches) - 1:
                end_pos = matches[i+1].start()
            else:
                end_pos = len(texto_completo)
            
            texto_declaracion = texto_completo[start_pos:end_pos]
            
            datos = self.extraer_datos_declaracion_individual(texto_declaracion, numero_formulario)
            if datos:
                declaraciones.append(datos)
        
        return declaraciones

    def extraer_datos_declaracion_individual(self, texto, numero_formulario):
        """Extrae datos de una declaración individual"""
        datos = {
            'Numero_Formulario_Declaracion': numero_formulario,
            'Archivo_PDF': os.path.basename(texto.split('\n')[0]) if texto else 'Desconocido'
        }
        
        for campo in self.CAMPOS_DI.values():
            if campo in self.patrones:
                valor = self.extraer_campo_individual(texto, self.patrones[campo], campo)
                if any(palabra in campo for palabra in ['Fecha', 'Aceptación', 'Levante']):
                    valor = self.normalizar_fecha_dd_mm_aaaa(valor, es_fecha=True)
                datos[campo] = valor
        
        return datos

    def extraer_campo_individual(self, texto, patrones, campo_nombre=""):
        """Extrae un campo específico de una declaración individual"""
        for patron in patrones:
            try:
                match = re.search(patron, texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                if match:
                    if match.groups():
                        for group_val in match.groups():
                            if group_val and group_val.strip():
                                return group_val.strip()
                    else:
                        return match.group(0).strip()
            except Exception:
                continue
        return "NO ENCONTRADO"

    def normalizar_fecha_dd_mm_aaaa(self, fecha_str, es_fecha=True):
        """Normaliza formato de fecha a dd-mm-aaaa"""
        if not fecha_str or fecha_str == "NO ENCONTRADO" or str(fecha_str).strip() == "":
            return "NO ENCONTRADO"
        
        if not es_fecha:
            return str(fecha_str).strip()
        
        try:
            if isinstance(fecha_str, datetime):
                return fecha_str.strftime('%d-%m-%Y')
            
            fecha_limpia = str(fecha_str).strip()
            fecha_limpia = re.sub(r'\s+', '', fecha_limpia)
            
            if len(fecha_limpia) > 10 and fecha_limpia.isdigit():
                return fecha_limpia
            
            patrones_fecha = [
                (r'^(\d{4})(\d{2})(\d{2})$', '%Y%m%d'),
                (r'(\d{4})-(\d{1,2})-(\d{1,2})', '%Y-%m-%d'),
                (r'(\d{4})/(\d{1,2})/(\d{1,2})', '%Y/%m/%d'),
                (r'(\d{1,2})-(\d{1,2})-(\d{4})', '%d-%m-%Y'),
                (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%d/%m/%Y'),
            ]
            
            for patron, formato in patrones_fecha:
                match = re.match(patron, fecha_limpia)
                if match:
                    try:
                        fecha_obj = datetime.strptime(fecha_limpia, formato)
                        return fecha_obj.strftime('%d-%m-%Y')
                    except ValueError:
                        continue
            
            return fecha_limpia
            
        except Exception as e:
            return str(fecha_str)

    def _normalizar_factura(self, factura_str):
        """Normaliza número de factura para comparación"""
        if not factura_str or factura_str == "NO ENCONTRADO":
            return ""
        
        factura_limpia = str(factura_str).strip().upper()
        factura_limpia = re.sub(r'\s*/\s*', '/', factura_limpia)
        factura_limpia = factura_limpia.replace(' ', '')
        factura_limpia = re.sub(r'[^\w\/\-]', '', factura_limpia)
        
        return factura_limpia

    def _emparejar_facturas_completo(self, facturas_declaraciones, facturas_formulario):
        """Empareja facturas manejando todos los casos posibles - SIN LOGS REPETITIVOS"""
        emparejamientos = {}
        
        # Crear conjuntos normalizados para comparación
        facturas_decl_norm = {self._normalizar_factura(f): f for f in facturas_declaraciones.values()}
        facturas_form_norm = {self._normalizar_factura(f): f for f in facturas_formulario}
        
        # CASO 1: Solo hay UNA factura en el formulario
        if len(facturas_formulario) == 1:
            factura_unica = facturas_formulario[0]
            
            # Esta misma factura se usa para TODAS las declaraciones
            for di_num, factura_decl in facturas_declaraciones.items():
                emparejamientos[di_num] = factura_unica
        
        # CASO 2: Hay múltiples facturas en el formulario
        else:
            facturas_disponibles = facturas_formulario.copy()
            
            # PRIMERO: Emparejamiento exacto
            for di_num, factura_decl in facturas_declaraciones.items():
                if di_num in emparejamientos:
                    continue
                    
                factura_decl_norm = self._normalizar_factura(factura_decl)
                
                for factura_form in facturas_disponibles[:]:
                    factura_form_norm = self._normalizar_factura(factura_form)
                    
                    if factura_decl_norm == factura_form_norm:
                        emparejamientos[di_num] = factura_form
                        facturas_disponibles.remove(factura_form)
                        break
            
            # SEGUNDO: Emparejamiento por parte numérica para los restantes
            for di_num, factura_decl in facturas_declaraciones.items():
                if di_num in emparejamientos:
                    continue
                    
                factura_decl_norm = self._normalizar_factura(factura_decl)
                
                if '/' in factura_decl_norm:
                    parte_principal_decl = factura_decl_norm.split('/')[0]
                else:
                    parte_principal_decl = factura_decl_norm
                
                for factura_form in facturas_disponibles[:]:
                    factura_form_norm = self._normalizar_factura(factura_form)
                    
                    if '/' in factura_form_norm:
                        parte_principal_form = factura_form_norm.split('/')[0]
                    else:
                        parte_principal_form = factura_form_norm
                    
                    if parte_principal_decl == parte_principal_form:
                        emparejamientos[di_num] = factura_form
                        facturas_disponibles.remove(factura_form)
                        break
            
            # TERCERO: Si aún quedan DI sin emparejar, usar las facturas restantes en orden
            di_sin_emparejar = [di for di in facturas_declaraciones.keys() if di not in emparejamientos]
            if di_sin_emparejar and facturas_disponibles:
                for i, di_num in enumerate(di_sin_emparejar):
                    if i < len(facturas_disponibles):
                        emparejamientos[di_num] = facturas_disponibles[i]
        
        # Verificar que todas las DI tengan una factura asignada
        for di_num in facturas_declaraciones.keys():
            if di_num not in emparejamientos:
                # Si no se pudo emparejar, usar la primera factura disponible
                if facturas_formulario:
                    emparejamientos[di_num] = facturas_formulario[0]
                else:
                    emparejamientos[di_num] = "NO ENCONTRADO"
        
        return emparejamientos

    def _comparar_nombres_optimizado(self, nombre_pdf, nombre_excel):
        """Función optimizada para comparación de nombres con cache"""
        cache_key = f"{nombre_pdf}_{nombre_excel}"
        if cache_key in self._cache_nombres:
            return self._cache_nombres[cache_key]
        
        resultado = self.corrector_nombres.comparar_por_letras(nombre_pdf, nombre_excel)
        self._cache_nombres[cache_key] = resultado
        return resultado

    def validar_campos_por_declaracion(self, datos_declaracion, anexos_formulario):
        """Valida los campos para una declaración específica con corrección automática de nombres - OPTIMIZADA"""
        resultados_validacion = []
        
        if anexos_formulario.empty and not (self.nit_proveedor and self.nombre_proveedor):
            return pd.DataFrame()
        
        numero_di = datos_declaracion.get('Numero_Formulario_Declaracion', 'NO ENCONTRADO')
        
        # Pre-calcular valores para evitar llamadas repetidas
        valor_nombre_pdf = datos_declaracion.get("11. Apellidos y Nombres / Razón Social Importador", "NO ENCONTRADO")
        nombre_corregido = None
        
        for campo_declaracion, config in self.MAPEOS_VALIDACION.items():
            resultado = {
                'Campos DI a Validar': campo_declaracion,
                'Datos Declaración': 'NO ENCONTRADO',
                'Datos Formulario': 'NO ENCONTRADO',
                'Numero DI': numero_di,
                'Coincidencias': '❌ NO COINCIDE'
            }
            
            try:
                codigo_esperado = config["codigo_formulario"]
                tipo = config["tipo"]
                cambia_por_declaracion = config["cambia_por_declaracion"]
                
                valor_declaracion = datos_declaracion.get(campo_declaracion, "NO ENCONTRADO")
                
                if tipo == "fecha" and valor_declaracion != "NO ENCONTRADO":
                    valor_declaracion = self.normalizar_fecha_dd_mm_aaaa(valor_declaracion, es_fecha=True)
                
                # ✅ APLICAR CORRECCIÓN DE NOMBRES PARA EL CAMPO 11 - OPTIMIZADO
                if campo_declaracion == "11. Apellidos y Nombres / Razón Social Importador":
                    if nombre_corregido is None:
                        nombre_corregido = self.corrector_nombres.corregir_nombre(
                            valor_nombre_pdf, 
                            self.nombre_proveedor if self.nombre_proveedor else ""
                        )
                    resultado['Datos Declaración'] = nombre_corregido
                else:
                    resultado['Datos Declaración'] = valor_declaracion
                
                if codigo_esperado == "PROVEEDOR":
                    if campo_declaracion == "5. Número de Identificación Tributaria (NIT)":
                        resultado['Datos Formulario'] = self.nit_proveedor if self.nit_proveedor else 'NO ENCONTRADO'
                        coincide = str(valor_declaracion).strip() == str(self.nit_proveedor).strip()
                        resultado['Coincidencias'] = '✅ COINCIDE' if coincide else '❌ NO COINCIDE'
                    
                    elif campo_declaracion == "11. Apellidos y Nombres / Razón Social Importador":
                        resultado['Datos Formulario'] = self.nombre_proveedor if self.nombre_proveedor else 'NO ENCONTRADO'
                        coincide = self._comparar_nombres_optimizado(
                            valor_nombre_pdf, 
                            self.nombre_proveedor if self.nombre_proveedor else ""
                        )
                        resultado['Coincidencias'] = '✅ COINCIDE' if coincide else '❌ NO COINCIDE'
                
                else:
                    anexos_filtrados = anexos_formulario[anexos_formulario['Codigo'] == codigo_esperado]
                    
                    if not anexos_filtrados.empty:
                        if cambia_por_declaracion:
                            if numero_di in self.facturas_emparejadas:
                                factura_emparejada = self.facturas_emparejadas[numero_di]
                                anexos_filtrados = anexos_filtrados[anexos_filtrados['Documento'] == factura_emparejada]
                        
                        if not anexos_filtrados.empty:
                            anexo_seleccionado = anexos_filtrados.iloc[0]
                            
                            if tipo == "documento":
                                valor_formulario = anexo_seleccionado['Documento']
                            else:
                                valor_formulario = anexo_seleccionado['Fecha']
                            
                            if valor_formulario and valor_formulario != "NO ENCONTRADO":
                                if tipo == "fecha":
                                    valor_formulario = self.normalizar_fecha_dd_mm_aaaa(valor_formulario, es_fecha=True)
                                
                                resultado['Datos Formulario'] = valor_formulario
                                
                                if tipo == "fecha":
                                    coincide = valor_declaracion == valor_formulario
                                else:
                                    coincide = str(valor_declaracion).strip() == str(valor_formulario).strip()
                                
                                resultado['Coincidencias'] = '✅ COINCIDE' if coincide else '❌ NO COINCIDE'
                                
                                if coincide:
                                    anexos_formulario.loc[anexo_seleccionado.name, 'Usado'] = True
                            else:
                                resultado['Datos Formulario'] = 'NO ENCONTRADO'
                                resultado['Coincidencias'] = '❌ NO COINCIDE'
                        else:
                            resultado['Datos Formulario'] = 'NO ENCONTRADO'
                            resultado['Coincidencias'] = '❌ NO COINCIDE'
                    else:
                        resultado['Datos Formulario'] = 'NO ENCONTRADO'
                        resultado['Coincidencias'] = '❌ NO COINCIDE'
                
                resultados_validacion.append(resultado)
                
            except Exception as e:
                st.error(f"❌ ERROR validando {campo_declaracion}: {e}")
                resultados_validacion.append(resultado)
        
        return pd.DataFrame(resultados_validacion)

    def procesar_validacion_completa(self, carpeta_pdf, carpeta_formulario):
        """Procesa la validación completa con manejo de múltiples declaraciones"""
        st.info("🚀 INICIANDO VALIDACIÓN COMPLETA...")
        
        # Buscar archivos
        archivos_pdf = glob.glob(os.path.join(carpeta_pdf, "*.pdf"))
        archivo_formulario = self.buscar_archivo_formulario(carpeta_formulario)
        
        if not archivos_pdf:
            st.error("❌ No se encontraron archivos PDF en la carpeta")
            return None, None, None
        
        if not archivo_formulario:
            st.error("❌ No se encontró archivo del formulario FMM")
            return None, None, None
        
        # Extraer información del formulario
        if not self.extraer_proveedor_formulario(archivo_formulario):
            st.error("❌ No se pudo extraer información del proveedor")
            return None, None, None
        
        anexos_formulario = self.extraer_anexos_formulario_robusto(archivo_formulario)
        if anexos_formulario.empty:
            st.error("❌ No se pudieron extraer anexos del formulario")
            return None, None, None
        
        # Extraer facturas del formulario
        facturas_formulario = anexos_formulario[
            anexos_formulario['Codigo'] == 6
        ]['Documento'].dropna().unique().tolist()
        
        # Procesar todas las declaraciones del PDF
        todas_declaraciones = []
        for archivo_pdf in archivos_pdf:
            declaraciones_pdf = self.extraer_todas_declaraciones_pdf(archivo_pdf)
            todas_declaraciones.extend(declaraciones_pdf)
        
        if not todas_declaraciones:
            st.error("❌ No se pudieron extraer declaraciones del PDF")
            return None, None, None
        
        st.success(f"✅ {len(todas_declaraciones)} declaraciones encontradas en el PDF")
        
        # Extraer facturas de las declaraciones
        facturas_declaraciones = {}
        for declaracion in todas_declaraciones:
            numero_di = declaracion.get('Numero_Formulario_Declaracion', 'DESCONOCIDO')
            factura = declaracion.get("51. No. Factura Comercial", "NO ENCONTRADO")
            if factura != "NO ENCONTRADO":
                facturas_declaraciones[numero_di] = factura
        
        # Emparejar facturas
        self.facturas_emparejadas = self._emparejar_facturas_completo(
            facturas_declaraciones, 
            facturas_formulario
        )
        
        # Validar cada declaración
        resultados_totales = []
        for declaracion in todas_declaraciones:
            numero_di = declaracion.get('Numero_Formulario_Declaracion', 'DESCONOCIDO')
            st.info(f"🔍 Validando declaración: {numero_di}")
            
            resultado_declaracion = self.validar_campos_por_declaracion(declaracion, anexos_formulario)
            if not resultado_declaracion.empty:
                resultados_totales.append(resultado_declaracion)
        
        if resultados_totales:
            df_resultados = pd.concat(resultados_totales, ignore_index=True)
            
            # Calcular estadísticas
            total_validaciones = len(df_resultados)
            coincidencias = len(df_resultados[df_resultados['Coincidencias'] == '✅ COINCIDE'])
            no_coincidencias = len(df_resultados[df_resultados['Coincidencias'] == '❌ NO COINCIDE'])
            
            st.success("📊 RESUMEN DE VALIDACIÓN:")
            st.success(f"   • Total de validaciones: {total_validaciones}")
            st.success(f"   • ✅ Coincidencias: {coincidencias}")
            st.success(f"   • ❌ No coincidencias: {no_coincidencias}")
            st.success(f"   • 📈 Porcentaje de éxito: {(coincidencias/total_validaciones)*100:.1f}%")
            
            return df_resultados, todas_declaraciones, anexos_formulario
        else:
            st.error("❌ No se generaron resultados de validación")
            return None, None, None

# =============================================================================
# CLASE PARA VALIDACIÓN DE FACTURAS (Del segundo código)
# =============================================================================

class ValidadorFacturas:
    def __init__(self):
        self.campos_requeridos = [
            "Número de factura",
            "Fecha de factura", 
            "Proveedor",
            "NIT",
            "Descripción",
            "Cantidad",
            "Precio unitario",
            "Valor total"
        ]
        
    def validar_estructura_excel(self, archivo_excel):
        """Valida la estructura básica del archivo Excel"""
        try:
            df = pd.read_excel(archivo_excel)
            
            # Verificar columnas requeridas
            columnas_faltantes = []
            for campo in self.campos_requeridos:
                if campo not in df.columns:
                    columnas_faltantes.append(campo)
            
            if columnas_faltantes:
                return False, f"Columnas faltantes: {', '.join(columnas_faltantes)}"
            
            # Verificar que haya datos
            if df.empty:
                return False, "El archivo no contiene datos"
            
            return True, "Estructura válida"
            
        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"
    
    def validar_formato_facturas(self, archivo_excel):
        """Valida el formato de las facturas en el Excel"""
        try:
            df = pd.read_excel(archivo_excel)
            resultados = []
            
            for idx, fila in df.iterrows():
                errores_fila = []
                
                # Validar número de factura
                if pd.isna(fila["Número de factura"]) or str(fila["Número de factura"]).strip() == "":
                    errores_fila.append("Número de factura vacío")
                
                # Validar fecha
                if pd.isna(fila["Fecha de factura"]):
                    errores_fila.append("Fecha de factura vacía")
                
                # Validar NIT
                if pd.isna(fila["NIT"]) or str(fila["NIT"]).strip() == "":
                    errores_fila.append("NIT vacío")
                
                # Validar valores numéricos
                if pd.isna(fila["Cantidad"]) or fila["Cantidad"] <= 0:
                    errores_fila.append("Cantidad inválida")
                
                if pd.isna(fila["Precio unitario"]) or fila["Precio unitario"] < 0:
                    errores_fila.append("Precio unitario inválido")
                
                if pd.isna(fila["Valor total"]) or fila["Valor total"] < 0:
                    errores_fila.append("Valor total inválido")
                
                # Validar cálculo
                if not pd.isna(fila["Cantidad"]) and not pd.isna(fila["Precio unitario"]):
                    calculado = fila["Cantidad"] * fila["Precio unitario"]
                    if abs(calculado - fila["Valor total"]) > 0.01:
                        errores_fila.append(f"Valor total no coincide: {calculado} vs {fila['Valor total']}")
                
                resultados.append({
                    "Línea": idx + 2,  # +2 porque Excel empieza en 1 y tiene headers
                    "Número de factura": fila["Número de factura"],
                    "Errores": ", ".join(errores_fila) if errores_fila else "✅ Válida",
                    "Estado": "✅ Válida" if not errores_fila else "❌ Con errores"
                })
            
            return pd.DataFrame(resultados)
            
        except Exception as e:
            st.error(f"Error en validación de facturas: {str(e)}")
            return pd.DataFrame()

# =============================================================================
# INTERFAZ STREAMLIT UNIFICADA
# =============================================================================

def main():
    # Configuración de estilo
    st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #2e86ab;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header principal
    st.markdown('<h1 class="main-header">📊 Dashboard de Validación de Importaciones</h1>', unsafe_allow_html=True)
    
    # Sidebar para navegación
    st.sidebar.title("🔧 Navegación")
    opcion = st.sidebar.radio(
        "Selecciona el módulo:",
        ["🏠 Inicio", "📋 Validación DIM vs Anexos", "🧾 Validación de Facturas", "📈 Resultados"]
    )
    
    # Inicializar validadores
    if 'validador_dim' not in st.session_state:
        st.session_state.validador_dim = ValidadorDeclaracionImportacionCompleto()
    
    if 'validador_facturas' not in st.session_state:
        st.session_state.validador_facturas = ValidadorFacturas()
    
    # Módulo de Inicio
    if opcion == "🏠 Inicio":
        st.markdown("""
        ## 🎯 Bienvenido al Sistema de Validación de Importaciones
        
        Este dashboard integrado permite:
        
        ### 📋 Validación DIM vs Anexos FMM
        - Comparación automática entre Declaraciones de Importación (DIM) y formularios FMM
        - Validación de campos críticos: NIT, nombres, facturas, manifiestos, etc.
        - Corrección automática de nombres por número de letras
        - Detección de inconsistencias y duplicados
        
        ### 🧾 Validación de Facturas  
        - Verificación de estructura de archivos Excel de facturas
        - Validación de formatos y cálculos
        - Detección de errores en datos de facturación
        
        ### 📈 Resultados
        - Reportes consolidados de validaciones
        - Métricas y estadísticas
        - Exportación de resultados
        
        ### 🚀 Comenzar
        1. Selecciona un módulo en la barra lateral
        2. Sube los archivos requeridos
        3. Ejecuta la validación
        4. Revisa los resultados y reportes
        """)
    
    # Módulo de Validación DIM vs Anexos
    elif opcion == "📋 Validación DIM vs Anexos":
        st.markdown('<h2 class="sub-header">📋 Validación DIM vs Anexos FMM</h2>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📄 Archivos PDF (Declaraciones)")
            pdf_files = st.file_uploader(
                "Selecciona archivos PDF de declaraciones",
                type=['pdf'],
                accept_multiple_files=True,
                key="pdf_uploader"
            )
        
        with col2:
            st.subheader("📊 Archivo Excel (Formulario FMM)")
            excel_file = st.file_uploader(
                "Selecciona archivo Excel del formulario FMM",
                type=['xlsx', 'xls'],
                key="excel_uploader"
            )
        
        if st.button("🚀 Ejecutar Validación DIM vs Anexos", type="primary"):
            if not pdf_files or not excel_file:
                st.error("❌ Por favor sube ambos tipos de archivos")
                return
            
            with st.spinner("🔄 Procesando validación..."):
                # Crear carpetas temporales
                with tempfile.TemporaryDirectory() as temp_dir:
                    pdf_dir = os.path.join(temp_dir, "pdfs")
                    excel_dir = os.path.join(temp_dir, "excel")
                    os.makedirs(pdf_dir, exist_ok=True)
                    os.makedirs(excel_dir, exist_ok=True)
                    
                    # Guardar archivos PDF
                    for pdf_file in pdf_files:
                        pdf_path = os.path.join(pdf_dir, pdf_file.name)
                        with open(pdf_path, "wb") as f:
                            f.write(pdf_file.getbuffer())
                    
                    # Guardar archivo Excel
                    excel_path = os.path.join(excel_dir, excel_file.name)
                    with open(excel_path, "wb") as f:
                        f.write(excel_file.getbuffer())
                    
                    # Ejecutar validación
                    resultados, declaraciones, anexos = st.session_state.validador_dim.procesar_validacion_completa(
                        pdf_dir, excel_dir
                    )
                    
                    # Guardar resultados en session state
                    st.session_state.resultados_dim = resultados
                    st.session_state.declaraciones_dim = declaraciones
                    st.session_state.anexos_dim = anexos
                    
                    if resultados is not None:
                        st.success("✅ Validación completada exitosamente!")
                        
                        # Mostrar resumen
                        st.subheader("📊 Resumen de Validación")
                        total = len(resultados)
                        coincidencias = len(resultados[resultados['Coincidencias'] == '✅ COINCIDE'])
                        porcentaje = (coincidencias / total) * 100 if total > 0 else 0
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total Validaciones", total)
                        col2.metric("Coincidencias", coincidencias)
                        col3.metric("Porcentaje Éxito", f"{porcentaje:.1f}%")
                        
                        # Mostrar tabla de resultados
                        st.subheader("📋 Resultados Detallados")
                        st.dataframe(resultados, use_container_width=True)
                        
                        # Botón de descarga
                        csv = resultados.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="📥 Descargar Resultados (CSV)",
                            data=csv,
                            file_name=f"resultados_validacion_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                            mime="text/csv"
                        )
    
    # Módulo de Validación de Facturas
    elif opcion == "🧾 Validación de Facturas":
        st.markdown('<h2 class="sub-header">🧾 Validación de Facturas</h2>', unsafe_allow_html=True)
        
        st.subheader("📊 Subir Archivo de Facturas")
        facturas_file = st.file_uploader(
            "Selecciona archivo Excel de facturas",
            type=['xlsx', 'xls'],
            key="facturas_uploader"
        )
        
        if st.button("🔍 Validar Facturas", type="primary"):
            if not facturas_file:
                st.error("❌ Por favor sube un archivo Excel de facturas")
                return
            
            with st.spinner("🔄 Validando facturas..."):
                # Validar estructura
                es_valido, mensaje = st.session_state.validador_facturas.validar_estructura_excel(facturas_file)
                
                if not es_valido:
                    st.error(f"❌ {mensaje}")
                    return
                
                st.success("✅ Estructura del archivo válida")
                
                # Validar formato de facturas
                resultados_facturas = st.session_state.validador_facturas.validar_formato_facturas(facturas_file)
                
                # Guardar en session state
                st.session_state.resultados_facturas = resultados_facturas
                
                if not resultados_facturas.empty:
                    st.subheader("📋 Resultados de Validación de Facturas")
                    
                    # Métricas
                    total_facturas = len(resultados_facturas)
                    facturas_validas = len(resultados_facturas[resultados_facturas['Estado'] == '✅ Válida'])
                    facturas_con_errores = total_facturas - facturas_validas
                    
                    col1, col2 = st.columns(2)
                    col1.metric("Total Facturas", total_facturas)
                    col2.metric("Facturas Válidas", facturas_validas)
                    
                    # Mostrar tabla
                    st.dataframe(resultados_facturas, use_container_width=True)
                    
                    # Botón de descarga
                    csv_facturas = resultados_facturas.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 Descargar Resultados Facturas (CSV)",
                        data=csv_facturas,
                        file_name=f"resultados_facturas_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv"
                    )
    
    # Módulo de Resultados
    elif opcion == "📈 Resultados":
        st.markdown('<h2 class="sub-header">📈 Resultados Consolidados</h2>', unsafe_allow_html=True)
        
        # Resultados DIM vs Anexos
        if 'resultados_dim' in st.session_state:
            st.subheader("📋 Resultados Validación DIM vs Anexos")
            
            resultados = st.session_state.resultados_dim
            total = len(resultados)
            coincidencias = len(resultados[resultados['Coincidencias'] == '✅ COINCIDE'])
            porcentaje = (coincidencias / total) * 100 if total > 0 else 0
            
            # Gráfico de resultados
            col1, col2 = st.columns([2, 1])
            
            with col1:
                chart_data = pd.DataFrame({
                    'Tipo': ['Coincidencias', 'No Coincidencias'],
                    'Cantidad': [coincidencias, total - coincidencias]
                })
                st.bar_chart(chart_data.set_index('Tipo'))
            
            with col2:
                st.metric("Porcentaje Éxito", f"{porcentaje:.1f}%")
                st.metric("Coincidencias", coincidencias)
                st.metric("No Coincidencias", total - coincidencias)
            
            # Tabla detallada
            st.dataframe(resultados, use_container_width=True)
        
        else:
            st.info("ℹ️ No hay resultados de validación DIM vs Anexos disponibles")
        
        st.markdown("---")
        
        # Resultados Facturas
        if 'resultados_facturas' in st.session_state:
            st.subheader("🧾 Resultados Validación de Facturas")
            
            resultados_facturas = st.session_state.resultados_facturas
            total_facturas = len(resultados_facturas)
            facturas_validas = len(resultados_facturas[resultados_facturas['Estado'] == '✅ Válida'])
            
            # Gráfico de resultados
            col1, col2 = st.columns([2, 1])
            
            with col1:
                chart_data_facturas = pd.DataFrame({
                    'Tipo': ['Válidas', 'Con Errores'],
                    'Cantidad': [facturas_validas, total_facturas - facturas_validas]
                })
                st.bar_chart(chart_data_facturas.set_index('Tipo'))
            
            with col2:
                st.metric("Total Facturas", total_facturas)
                st.metric("Facturas Válidas", facturas_validas)
                st.metric("Facturas con Errores", total_facturas - facturas_validas)
            
            # Tabla detallada
            st.dataframe(resultados_facturas, use_container_width=True)
        
        else:
            st.info("ℹ️ No hay resultados de validación de facturas disponibles")
        
        # Reporte consolidado
        if st.button("📊 Generar Reporte Consolidado"):
            st.subheader("📊 Reporte Consolidado")
            
            # Aquí podrías generar un reporte combinando ambos tipos de validación
            st.success("✅ Reporte consolidado generado (funcionalidad en desarrollo)")
            
            # Ejemplo de métricas combinadas
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if 'resultados_dim' in st.session_state:
                    st.metric("Validaciones DIM", len(st.session_state.resultados_dim))
                else:
                    st.metric("Validaciones DIM", 0)
            
            with col2:
                if 'resultados_facturas' in st.session_state:
                    st.metric("Facturas Validadas", len(st.session_state.resultados_facturas))
                else:
                    st.metric("Facturas Validadas", 0)
            
            with col3:
                st.metric("Procesos Completados", 
                         int('resultados_dim' in st.session_state) + 
                         int('resultados_facturas' in st.session_state))

if __name__ == "__main__":
    main()