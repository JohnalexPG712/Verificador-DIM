import pandas as pd
import pdfplumber
import re
import os
import glob
from datetime import datetime
import numpy as np
from collections import OrderedDict
from openpyxl import load_workbook
import warnings
import unicodedata

# =============================================================================
# CLASE PARA CORRECCIÓN DE NOMBRES
# =============================================================================

class CorrectorNombres:
    """Clase simplificada para corregir nombres basada en número de letras"""
    
    def normalizar_texto(self, texto):
        if not texto or texto == "NO ENCONTRADO":
            return ""
        texto = str(texto).upper()
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto) 
                       if unicodedata.category(c) != 'Mn')
        texto = re.sub(r'[^A-Z]', '', texto)
        return texto
    
    def comparar_por_letras(self, nombre_pdf, nombre_excel):
        if not nombre_pdf or not nombre_excel:
            return False
        pdf_normalizado = self.normalizar_texto(nombre_pdf)
        excel_normalizado = self.normalizar_texto(nombre_excel)
        return len(pdf_normalizado) == len(excel_normalizado)
    
    def corregir_nombre(self, nombre_pdf, nombre_excel):
        if not nombre_excel or nombre_excel == "NO ENCONTRADO":
            return nombre_pdf
        if self.comparar_por_letras(nombre_pdf, nombre_excel):
            return nombre_excel
        else:
            return nombre_pdf

# =============================================================================
# CLASE 1: EXTRACCIÓN DE PDFs (DIAN)
# =============================================================================

class ExtractorDIANSimplificado:
    def __init__(self):
        self.CAMPOS_DI = {
            "4.": "4. Número DI",
            "55.": "55. Cod. de Bandera", 
            "58.": "58. Tasa de Cambio",
            "59.": "59. Subpartida Arancelaria",
            "62.": "62. Cod. Modalidad",
            "66.": "66. Cod. Pais de Origen",
            "70.": "70. Cod. Pais Compra",
            "71.": "71. Peso Bruto kgs.",
            "72.": "72. Peso Neto kgs.",
            "74.": "74. Número de Bultos",
            "77.": "77. Cantidad dcms.",
            "78.": "78. Valor FOB USD",
            "79.": "79. Valor Fletes USD",
            "80.": "80. Valor Seguros USD",
            "81.": "81. Valor Otros Gastos USD"
        }

        self.patrones = {
            "4. Número DI": [
                r"(?:^|\n)\s*4\s*\.?\s*N[úu]mero\s*de\s*formulario[\s\S]*?(\d{15,16})",
                r"(?:^|\n)\s*4\s*\.?\s*N[úu]mero\s*de\s*formulario[\s\S]*?([\d\-]{15})",
                r"4\s*\.?\s*N[úu]mero\s*de\s*formulario\s*[:\-]?\s*(\d{17}(?:-\d)?)"
            ],
            "55. Cod. de Bandera": [r"55\s*\.\s*?C[oó]digo\s*de.*?\n(?:\s*\d+\s+){2}(\d+)"],
            "58. Tasa de Cambio": [r"58\s*\.?\s*Tasa\s*de\s*cambio\b(?:\s*\$?\s*cvs\.?)?[\s\S]{0,200}?([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2}))"],
            "59. Subpartida Arancelaria": [
                r"59\s*\.?\s*Subpartida\s*arancelaria\s*\d+\s*\.\s*Cod\s*\.\s*\d+\s*\.\s*Cod\s*\.\s*\d+\s*\.\s*Cod\s*\.\s*Modalidad\s*\d+\s*\.\s*No\s*\.\s*cuotas\s*\d+\s*\.\s*Valor\s*cuota\s*USD\s*\d+\s*\.\s*Periodicidad\s*del\s*\d+\s*\.\s*Cod\s*\.\s*país\s*\d+\s*\.\s*Cod\s*\.\s*Acuerdo\s*([\d]{10})",
                r"59\s*\.?\s*Subpartida\s*arancelaria[\s\S]{0,150}?\b(\d{10})\b"
            ],
            "62. Cod. Modalidad": [r"62\s*\.?\s*Cod\s*\.\s*Modalidad\s*(?:(?:.*?\n)|(?:(?:[:;.\-]|\s)+))[^\n]*?\b([A-Z]\d{3})\b"],
            "66. Cod. Pais de Origen": [r"66\s*\.?\s*Cod\s*\.\s*país[\s\S]*?\n.*?\n.*?\b(\d{3})\b"],
            "70. Cod. Pais Compra": [r"70\s*\.?\s*Cod\s*\.\s*país[\s\S]*?\n.*?\n.*?\b(\d{3})\b"],
            "71. Peso Bruto kgs.": [r"71\s*\.?\s*Peso\s*bruto\s*kgs\s*\.?\s*dcms\s*\.?[\s\S]{0,500}?(\d{1,3}(?:\.\d{3})*\.\d{2})"],
            "72. Peso Neto kgs.": [r"72\s*\.?\s*Peso\s*neto\s*kgs\s*\.?\s*dcms\s*\.?[\s\S]{0,500}?\d{1,3}(?:\.\d{3})*\.\d{2}[\s\S]{0,100}?(\d{1,3}(?:\.\d{3})*\.\d{2})"],
            "74. Número de Bultos": [
               r"74\s*\.\s*?\s*No\s*\.\s*bultos[\s\S]*?embalaje\s+(\d+[\.,]?\d*)",
               r"(?is)(?:embalaje[\s\S]{0,200}?\b[A-Z]{2,3}\b[\s\S]{0,50}?(\d{1,3}(?:\.\d{3})*)|embalaje[\s\S]{0,80}?(\d{1,3}(?:\.\d{3})*))" 
            ],
            "77. Cantidad dcms.": [r"77\s*\.?\s*Cantidad\s*dcms\.[\s\S]*?comercial\s+(\d{1,4}(?:\.\d{3})*\.\d{2})"],
            "78. Valor FOB USD": [r"78\s*\.?\s*Valor\s*FOB\s*USD[\s\S]*?\n\s*([\d.,]+)"],
            "79. Valor Fletes USD": [r"79\s*\.?\s*Valor\s*fletes\s*USD[\s\S]*?\n\s*[\d.,]+\s+([\d.,]+)"],
            "80. Valor Seguros USD": [r"80\s*\.?\s*Valor\s*Seguros\s*USD[\s\S]*?\n\s*([\d.,]+)"],
            "81. Valor Otros Gastos USD": [r"81\s*\.?\s*Valor\s*Otros\s*Gastos\s*USD[\s\S]*?\n\s*[\d.,]+\s+([\d.,]+)"]
        }

    def normalizar_numero_entero(self, numero_str, campo_nombre=""):
        if not isinstance(numero_str, str) or numero_str == "NO ENCONTRADO":
            return np.nan

        if '62. Cod. Modalidad' in campo_nombre:
            return numero_str.strip()

        if '74. Número de Bultos' in campo_nombre:
            # 1. Eliminar los puntos completamente (asumimos que son miles: 7.333 -> 7333)
            cleaned_str = numero_str.replace('.', '')
            # 2. Reemplazar comas por puntos (por si hubiera decimales reales: 7,5 -> 7.5)
            cleaned_str = cleaned_str.replace(',', '.')
            # 3. Limpiar cualquier basura restante
            cleaned_str = re.sub(r'[^\d.]', '', cleaned_str)
            try:
                valor = float(cleaned_str)
                return valor
            except ValueError:
                return np.nan          

        if any(codigo in campo_nombre for codigo in ['55. Cod. de Bandera', '66. Cod. Pais de Origen', '70. Cod. Pais Compra']):
            if numero_str.isdigit():
                return numero_str
            if ' - ' in numero_str:
                codigo = numero_str.split(' - ')[0].strip()
                if codigo.isdigit():
                    return codigo
            return numero_str

        cleaned_str = re.sub(r'(?<=\d)\.(?=\d{3})', '', numero_str)
        cleaned_str = re.sub(r'(?<=\d),(?=\d{3})', '', cleaned_str)
        if ',' in cleaned_str and '.' not in cleaned_str:
            cleaned_str = cleaned_str.replace(',', '.')

        try:
            valor = float(cleaned_str)
            if valor.is_integer():
                return int(valor)
            return valor
        except ValueError:
            return np.nan

    def extraer_texto_pdf(self, pdf_path):
        texto_completo = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text(x_tolerance=3, y_tolerance=3)
                    if texto:
                        texto_completo += texto + "\n\n"
            return texto_completo
        except Exception:
            return ""

    def extraer_campo(self, texto, patrones, campo_nombre=""):
        for patron in patrones:
            try:
                match = re.search(patron, texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                if match:
                    if match.groups():
                        for group_val in match.groups():
                            if group_val and group_val.strip():
                                return group_val.strip()
                    else:
                        return match.group(0).strip()
            except Exception:
                pass
        return "NO ENCONTRADO"

    def extraer_multiples_di_de_texto(self, texto_completo, pdf_filename):
        di_bloques = []
        form_number_matches = []
        for patron in self.patrones["4. Número DI"]:
            matches = list(re.finditer(patron, texto_completo, re.IGNORECASE | re.MULTILINE | re.DOTALL))
            form_number_matches.extend(matches)
        
        form_number_matches.sort(key=lambda m: m.start())
        
        if not form_number_matches:
            form_num = self.extraer_campo(texto_completo, self.patrones["4. Número DI"], "4. Número DI")
            if form_num != "NO ENCONTRADO":
                di_bloques.append({'form_number': form_num, 'text': texto_completo})
            else:
                di_bloques.append({'form_number': "Desconocido_" + pdf_filename, 'text': texto_completo})
            return di_bloques

        unique_matches = []
        last_end = -1
        for match in form_number_matches:
            if match.start() >= last_end:
                unique_matches.append(match)
                last_end = match.end()
        
        for i, match in enumerate(unique_matches):
            form_number = match.group(1).strip()
            start_index = match.start()
            end_index = unique_matches[i+1].start() if i + 1 < len(unique_matches) else len(texto_completo)
            di_text_block = texto_completo[start_index:end_index]
            di_bloques.append({'form_number': form_number, 'text': di_text_block})
        return di_bloques

    def procesar_di_individual(self, di_text_block, form_number, pdf_filename):
        resultados = OrderedDict()
        resultados['Nombre Archivo PDF'] = pdf_filename
        resultados["4. Número DI"] = form_number

        for _, nombre_campo in self.CAMPOS_DI.items():
            if nombre_campo == "4. Número DI":
                continue 
            if nombre_campo in self.patrones:
                valor = self.extraer_campo(di_text_block, self.patrones[nombre_campo], nombre_campo)
                resultados[nombre_campo] = self.normalizar_numero_entero(valor, nombre_campo)
            else:
                resultados[nombre_campo] = "PATRON NO CONFIGURADO"
        return resultados

    def procesar_multiples_dis(self, folder_path):
        if not os.path.isdir(folder_path): return None
        all_results = []
        pdf_files = glob.glob(os.path.join(folder_path, "*.pdf"))
        if not pdf_files: return None

        for pdf_file_path in pdf_files:
            pdf_filename = os.path.basename(pdf_file_path)
            texto_completo_pdf = self.extraer_texto_pdf(pdf_file_path)
            if not texto_completo_pdf: continue

            di_bloques = self.extraer_multiples_di_de_texto(texto_completo_pdf, pdf_filename)
            if not di_bloques: continue

            for bloque in di_bloques:
                resultados_di = self.procesar_di_individual(bloque['text'], bloque['form_number'], pdf_filename)
                if resultados_di:
                    all_results.append(resultados_di)
        return pd.DataFrame(all_results) if all_results else None

# =============================================================================
# CLASE 2: COMPARACIÓN DE DATOS
# =============================================================================

class ComparadorDatos:
    def __init__(self):
        self.campos_comparacion_individual = {
            'pais_origen': ('66. Cod. Pais de Origen', 'PAIS ORIGEN'),
            'pais_compra': ('70. Cod. Pais Compra', 'PAIS COMPRA'), 
            'bandera': ('55. Cod. de Bandera', 'BANDERA')
        }
        self.campos_consistencia = {
            '58. Tasa de Cambio': '58. Tasa de Cambio',
            '62. Cod. Modalidad': '62. Cod. Modalidad'
        }
        self.campos_bultos = {
            'numero_bultos': ('74. Número de Bultos', 'NUMERO BULTOS')
        }
        self.campos_acumulables = {
            'peso_bruto': ('71. Peso Bruto kgs.', 'PESO BRUTO'),
            'peso_neto': ('72. Peso Neto kgs.', 'PESO NETO'),
            'cantidad': ('77. Cantidad dcms.', 'CANTIDAD'),
            'valor_fob': ('78. Valor FOB USD', 'VALOR FOB'),
            'valor_fletes': ('79. Valor Fletes USD', 'VALOR FLETES'),
            'valor_seguro': ('80. Valor Seguros USD', 'VALOR SEGURO'),
            'otros_gastos': ('81. Valor Otros Gastos USD', 'OTROS GASTOS')
        }
        self.campos_subpartida_arancelaria = {
            'subpartida_arancelaria': ('59. Subpartida Arancelaria', 'SUBPARTIDA')
        }

    def es_valor_valido(self, valor):
        if pd.isna(valor) or valor in ["N/A", "NO ENCONTRADO", "", None]:
            return False
        return True

    def detectar_multiples_subpartidas(self, datos_subpartidas):
        if datos_subpartidas is None or datos_subpartidas.empty:
            return False
        if 'subpartida' in datos_subpartidas.columns:
            return datos_subpartidas['subpartida'].nunique() > 1
        return False

    def emparejar_di_con_subpartida(self, datos_dian, datos_subpartidas):
        emparejamientos = []
        if datos_subpartidas is None or datos_subpartidas.empty: return emparejamientos
        
        if self.detectar_multiples_subpartidas(datos_subpartidas):
            for _, di in datos_dian.iterrows():
                subpartida_di = di.get('59. Subpartida Arancelaria', 'NO ENCONTRADO')
                subpartida_correspondiente = None
                for _, subpartida in datos_subpartidas.iterrows():
                    subpartida_excel = subpartida.get('subpartida', 'NO ENCONTRADO')
                    if str(subpartida_di).strip() == str(subpartida_excel).strip():
                        subpartida_correspondiente = subpartida
                        break
                if subpartida_correspondiente is None and not datos_subpartidas.empty:
                    subpartida_correspondiente = datos_subpartidas.iloc[0]
                emparejamientos.append({'di': di, 'subpartida': subpartida_correspondiente})
        else:
            subpartida_unica = datos_subpartidas.iloc[0] if not datos_subpartidas.empty else None
            for _, di in datos_dian.iterrows():
                emparejamientos.append({'di': di, 'subpartida': subpartida_unica})
        return emparejamientos

    def calcular_totales_subpartidas_excel(self, datos_subpartidas):
        if datos_subpartidas is None or datos_subpartidas.empty: return {}
        totales = {}
        campos_acumulables_excel = [
            'peso_bruto', 'peso_neto', 'cantidad', 'valor_fob', 
            'valor_fletes', 'valor_seguro', 'otros_gastos', 'numero_bultos'
        ]
        for campo in campos_acumulables_excel:
            if campo in datos_subpartidas.columns:
                totales[campo] = datos_subpartidas[campo].sum()
            else:
                totales[campo] = 0
        return totales

    def calcular_totales_di(self, datos_dian):
        if datos_dian is None or datos_dian.empty: return {}
        totales = {}
        campos_acumulables_di = [
            '71. Peso Bruto kgs.', '72. Peso Neto kgs.', '77. Cantidad dcms.',
            '78. Valor FOB USD', '79. Valor Fletes USD', '80. Valor Seguros USD',
            '81. Valor Otros Gastos USD'
        ]
        for campo in campos_acumulables_di:
            if campo in datos_dian.columns:
                valores_validos = datos_dian[campo].apply(self.es_valor_valido)
                totales[campo] = datos_dian[valores_validos][campo].sum()
            else:
                totales[campo] = 0
        return totales

    def obtener_filas_validas_para_totales(self, datos_dian):
        if datos_dian is None or datos_dian.empty: return pd.DataFrame()
        campos_criticos = ['55. Cod. de Bandera', '66. Cod. Pais de Origen', '70. Cod. Pais Compra']
        mascara_valida = pd.Series([True] * len(datos_dian), index=datos_dian.index)
        for campo in campos_criticos:
            if campo in datos_dian.columns:
                mascara_campo = datos_dian[campo].apply(self.es_valor_valido)
                mascara_valida = mascara_valida & mascara_campo
        return datos_dian[mascara_valida]

    def formatear_numero_entero(self, valor, campo_nombre="", es_individual=True):
        if not self.es_valor_valido(valor): return "N/A"
        try:
            if '62. Cod. Modalidad' in campo_nombre: return str(valor)
            if '74. Número de Bultos' in campo_nombre:
                if isinstance(valor, (int, float)):
                    if valor.is_integer(): return int(valor)
                    else: return float(valor)
                elif isinstance(valor, str):
                    try:
                        cleaned = valor.replace(',', '.').strip()
                        if '.' in cleaned: return float(cleaned)
                        else: return int(cleaned)
                    except: return valor
                return valor
            if any(codigo in campo_nombre for codigo in ['55. Cod. de Bandera', '66. Cod. Pais de Origen', '70. Cod. Pais Compra']):
                if isinstance(valor, (int, float)): return f"{int(valor):03d}"
                elif isinstance(valor, str) and valor.isdigit(): return f"{int(valor):03d}"
                else: return str(valor)
            if isinstance(valor, (int, float)): return valor
            elif isinstance(valor, str):
                try:
                    cleaned = valor.replace(',', '.').strip()
                    if '.' in cleaned: return float(cleaned)
                    else: return int(cleaned)
                except: return valor
            return valor
        except: return valor

    def extraer_numero_pais(self, valor):
        if not self.es_valor_valido(valor): return None
        valor_str = str(valor).strip()
        if valor_str.isdigit(): return valor_str
        if ' - ' in valor_str:
            partes = valor_str.split(' - ')
            if partes and partes[0].strip().isdigit(): return partes[0].strip()
        numeros = re.findall(r'\d+', valor_str)
        if numeros: return numeros[0]
        return valor_str
        
    def comparar_valor_individual_critico(self, valor_dian, valor_subpartida, campo_nombre):
        valor_dian_formateado = self.formatear_numero_entero(valor_dian, campo_nombre, es_individual=True)
        if not self.es_valor_valido(valor_dian): return f"❌ {valor_dian_formateado}", False
        if not self.es_valor_valido(valor_subpartida): return f"✅ {valor_dian_formateado}", True
        
        numero_dian = self.extraer_numero_pais(valor_dian)
        numero_subpartida = self.extraer_numero_pais(valor_subpartida)
        
        if numero_dian and numero_subpartida and numero_dian == numero_subpartida:
            return f"✅ {valor_dian_formateado}", True
        else:
            return f"❌ {valor_dian_formateado}", False
    
    def verificar_consistencia_campo(self, datos_dian, campo_dian, numero_di):
        if campo_dian not in datos_dian.columns: return f"❌ NO ENCONTRADO"
        valor_actual = datos_dian[datos_dian["4. Número DI"] == numero_di][campo_dian].iloc[0] if not datos_dian[datos_dian["4. Número DI"] == numero_di].empty else "NO ENCONTRADO"
        
        if not self.es_valor_valido(valor_actual): return f"❌ N/A"
        valor_actual_formateado = self.formatear_numero_entero(valor_actual, campo_dian, es_individual=True)
        if campo_dian == "62. Cod. Modalidad": return valor_actual_formateado
            
        if campo_dian == "58. Tasa de Cambio":
            valores_unicos = datos_dian[campo_dian].apply(self.es_valor_valido)
            valores_validos = datos_dian[valores_unicos][campo_dian].unique()
            if len(valores_validos) == 1: return f"✅ {valor_actual_formateado}"
            else:
                valores_numericos = [v for v in valores_validos if isinstance(v, (int, float)) and not pd.isna(v)]
                if len(valores_numericos) > 1:
                    min_val = min(valores_numericos)
                    max_val = max(valores_numericos)
                    if (max_val - min_val) / min_val < 0.05: return f"✅ {valor_actual_formateado}"
                
                valor_mas_comun = datos_dian[campo_dian].mode().iloc[0] if not datos_dian[campo_dian].mode().empty else None
                if valor_actual != valor_mas_comun: return f"❌ {valor_actual_formateado}"
                else: return f"✅ {valor_actual_formateado}"
        return valor_actual_formateado       

    def determinar_resultado_final(self, fila_dian, fila_subpartida, multiples_subpartidas_excel=False):
        errores_criticos = False
        for campo, (campo_dian, campo_subpartida) in self.campos_comparacion_individual.items():
            valor_dian = fila_dian.get(campo_dian, "NO ENCONTRADO")
            valor_subpartida = fila_subpartida.get(campo, "NO ENCONTRADO")
            
            if self.es_valor_valido(valor_subpartida) and self.es_valor_valido(valor_dian):
                numero_dian = self.extraer_numero_pais(valor_dian)
                numero_subpartida = self.extraer_numero_pais(valor_subpartida)
                if numero_dian and numero_subpartida:
                    if numero_dian != numero_subpartida:
                        errores_criticos = True; break
                else:
                    if str(valor_dian).strip() != str(valor_subpartida).strip():
                        errores_criticos = True; break
            elif self.es_valor_valido(valor_dian) and not self.es_valor_valido(valor_subpartida):
                errores_criticos = True; break
            elif not self.es_valor_valido(valor_dian) and self.es_valor_valido(valor_subpartida):
                errores_criticos = True; break
            else:
                errores_criticos = True; break
        
        if not errores_criticos and multiples_subpartidas_excel:
            subpartida_dian = fila_dian.get("59. Subpartida Arancelaria", "NO ENCONTRADO")
            subpartida_excel = fila_subpartida.get("subpartida", "NO ENCONTRADO")
            if self.es_valor_valido(subpartida_dian) and self.es_valor_valido(subpartida_excel):
                if str(subpartida_dian).strip() != str(subpartida_excel).strip(): errores_criticos = True
            elif self.es_valor_valido(subpartida_dian) and not self.es_valor_valido(subpartida_excel): errores_criticos = True
            elif not self.es_valor_valido(subpartida_dian) and self.es_valor_valido(subpartida_excel): errores_criticos = True
        
        return errores_criticos

    def generar_reporte_tabular(self, datos_dian, datos_subpartidas):
        if datos_dian is None or datos_dian.empty or datos_subpartidas is None or datos_subpartidas.empty:
            return pd.DataFrame()
        
        multiples_subpartidas = self.detectar_multiples_subpartidas(datos_subpartidas)
        print(f"🔍 {'MÚLTIPLES SUBPARTIDAS' if multiples_subpartidas else 'SUBPARTIDA ÚNICA'} detectadas")
        
        emparejamientos = self.emparejar_di_con_subpartida(datos_dian, datos_subpartidas)
        reporte_filas = []
        
        for emparejamiento in emparejamientos:
            di = emparejamiento['di']
            subpartida = emparejamiento['subpartida']
            numero_di = di.get("4. Número DI", "Desconocido")
            # SILENCIADO: print(f"\n🔍 Procesando DI: {numero_di}")
            
            fila_reporte = {"4. Número DI": numero_di}
            
            for campo_consistencia, campo_dian in self.campos_consistencia.items():
                fila_reporte[campo_dian] = self.verificar_consistencia_campo(datos_dian, campo_dian, numero_di)
            
            for campo, (campo_dian, _) in self.campos_comparacion_individual.items():
                valor_dian = di.get(campo_dian, "NO ENCONTRADO")
                valor_subpartida = subpartida.get(campo, "NO ENCONTRADO") if subpartida is not None else "NO ENCONTRADO"
                valor_formateado, _ = self.comparar_valor_individual_critico(valor_dian, valor_subpartida, campo_dian)
                fila_reporte[f"{campo_dian} DI"] = valor_formateado
                fila_reporte[f"{campo_dian} Subpartida"] = self.formatear_numero_entero(valor_subpartida, f"{campo_dian} Subpartida", es_individual=True)
            
            for campo, (campo_dian, _) in self.campos_subpartida_arancelaria.items():
                valor_dian = di.get(campo_dian, "NO ENCONTRADO")
                valor_subpartida = subpartida.get('subpartida', "NO ENCONTRADO") if subpartida is not None else "NO ENCONTRADO"
                mostrar_emojis = multiples_subpartidas
                coinciden = self.es_valor_valido(valor_dian) and self.es_valor_valido(valor_subpartida) and str(valor_dian).strip() == str(valor_subpartida).strip()
                
                val_di_fmt = self.formatear_numero_entero(valor_dian, campo_dian)
                val_sub_fmt = self.formatear_numero_entero(valor_subpartida, f'{campo_dian} Subpartida')
                
                if mostrar_emojis:
                    emoji_di = "✅" if self.es_valor_valido(valor_dian) else "❌"
                    emoji_sub = "✅" if coinciden else "❌"
                    fila_reporte[f"{campo_dian} DI"] = f"{emoji_di} {val_di_fmt}"
                    fila_reporte[f"{campo_dian} Subpartida"] = f"{emoji_sub} {val_sub_fmt}"
                    # SILENCIADO: print(f"   📊 Subpartida - DI: {emoji_di} {val_di_fmt}, Excel: {emoji_sub} {val_sub_fmt}")
                else:
                    fila_reporte[f"{campo_dian} DI"] = val_di_fmt
                    fila_reporte[f"{campo_dian} Subpartida"] = val_sub_fmt

            for campo, (campo_dian, _) in self.campos_bultos.items():
                valor_dian = di.get(campo_dian, None)
                valor_subpartida = subpartida.get(campo, None) if subpartida is not None else None
                
                val_di_fmt = self.formatear_numero_entero(valor_dian, campo_dian) if self.es_valor_valido(valor_dian) else "N/A"
                val_sub_fmt = self.formatear_numero_entero(valor_subpartida, campo_dian) if self.es_valor_valido(valor_subpartida) else "N/A"
                
                fila_reporte[f"{campo_dian} DI"] = val_di_fmt if val_di_fmt != "N/A" else None
                fila_reporte[f"{campo_dian} Subpartida"] = val_sub_fmt if val_sub_fmt != "N/A" else None
                
                # SILENCIADO: print(f"   📦 Bultos - DI: {val_di_fmt}, Subpartida: {val_sub_fmt}")
                
            for campo, (campo_dian, _) in self.campos_acumulables.items():
                valor_dian = di.get(campo_dian, None)
                valor_subpartida = subpartida.get(campo, None) if subpartida is not None else None
                fila_reporte[f"{campo_dian} DI"] = valor_dian if self.es_valor_valido(valor_dian) else None
                fila_reporte[f"{campo_dian} Subpartida"] = valor_subpartida if self.es_valor_valido(valor_subpartida) else None
            
            tiene_errores = self.determinar_resultado_final(di, subpartida if subpartida is not None else {}, multiples_subpartidas)
            fila_reporte["Resultado verificación"] = "❌ CON DIFERENCIAS" if tiene_errores else "✅ CONFORME"
            reporte_filas.append(fila_reporte)
            # SILENCIADO: print(f"   {'❌' if tiene_errores else '✅'} DI: {numero_di} - {'CON DIFERENCIAS' if tiene_errores else 'CONFORME'}")
        
        if multiples_subpartidas:
            self._agregar_totales_multiples_subpartidas(reporte_filas, datos_dian, datos_subpartidas)
        else:
            self._agregar_totales_una_subpartida(reporte_filas, datos_dian, datos_subpartidas.iloc[0] if not datos_subpartidas.empty else {})
        
        df_reporte = pd.DataFrame(reporte_filas)
        columnas_ordenadas = self._ordenar_columnas_reporte_con_di(df_reporte)
        return df_reporte[columnas_ordenadas]

    def _agregar_totales_multiples_subpartidas(self, reporte_filas, datos_dian, datos_subpartidas):
        """Agrega fila de totales (Múltiples Subpartidas)"""
        totales_di = self.calcular_totales_di(datos_dian)
        totales_subpartidas = self.calcular_totales_subpartidas_excel(datos_subpartidas)
        fila_totales = {"4. Número DI": "VALORES ACUMULADOS (MÚLTIPLES SUBPARTIDAS)"}
        tiene_errores_totales = False
        
        for campo in self.campos_consistencia: fila_totales[campo] = "N/A"
        for campo, (campo_dian, _) in self.campos_comparacion_individual.items():
            fila_totales[f"{campo_dian} DI"] = "N/A"
            fila_totales[f"{campo_dian} Subpartida"] = "N/A"
        for campo, (campo_dian, _) in self.campos_subpartida_arancelaria.items():
            fila_totales[f"{campo_dian} DI"] = "N/A"
            fila_totales[f"{campo_dian} Subpartida"] = "N/A"
            
        valor_bultos_di_consolidado = 0
        if '74. Número de Bultos' in datos_dian.columns:
            s_bultos = pd.to_numeric(datos_dian['74. Número de Bultos'], errors='coerce')
            unicos = s_bultos.dropna().unique()
            if len(unicos) > 0:
                valor_bultos_di_consolidado = unicos[0]
        
        valor_bultos_excel_suma = totales_subpartidas.get('numero_bultos', 0)
        
        for campo, (campo_dian, _) in self.campos_bultos.items():
            nombre_campo_di = f"{campo_dian} DI"
            nombre_campo_subpartida = f"{campo_dian} Subpartida"
            
            try:
                diff = abs(valor_bultos_di_consolidado - valor_bultos_excel_suma)
                txt_di = f"{valor_bultos_di_consolidado:.0f}"
                txt_excel = f"{valor_bultos_excel_suma:.0f}"
                
                if diff < 1.0:
                    fila_totales[nombre_campo_di] = f"✅ {txt_di}"
                    fila_totales[nombre_campo_subpartida] = f"✅ {txt_excel}"
                else:
                    fila_totales[nombre_campo_di] = f"❌ {txt_di}"
                    fila_totales[nombre_campo_subpartida] = f"❌ {txt_excel}"
                    tiene_errores_totales = True
            except Exception as e:
                fila_totales[nombre_campo_di] = str(valor_bultos_di_consolidado)
                fila_totales[nombre_campo_subpartida] = str(valor_bultos_excel_suma)

        for campo, (campo_dian, _) in self.campos_acumulables.items():
            nombre_campo_di = f"{campo_dian} DI"
            nombre_campo_subpartida = f"{campo_dian} Subpartida"
            total_di = totales_di.get(campo_dian, 0)
            campo_subpartida = None
            if campo_dian == "71. Peso Bruto kgs.": campo_subpartida = "peso_bruto"
            elif campo_dian == "72. Peso Neto kgs.": campo_subpartida = "peso_neto"
            elif campo_dian == "77. Cantidad dcms.": campo_subpartida = "cantidad"
            elif campo_dian == "78. Valor FOB USD": campo_subpartida = "valor_fob"
            elif campo_dian == "79. Valor Fletes USD": campo_subpartida = "valor_fletes"
            elif campo_dian == "80. Valor Seguros USD": campo_subpartida = "valor_seguro"
            elif campo_dian == "81. Valor Otros Gastos USD": campo_subpartida = "otros_gastos"
            
            total_subpartida = totales_subpartidas.get(campo_subpartida, 0) if campo_subpartida else 0
            
            try:
                if total_di != 0 and total_subpartida != 0:
                    diferencia_absoluta = abs(float(total_di) - float(total_subpartida))
                    diferencia_porcentual = (diferencia_absoluta / float(total_subpartida)) * 100
                    coincide = False
                    if campo_dian == "78. Valor FOB USD": coincide = diferencia_absoluta < 1.0 or diferencia_porcentual < 1.0
                    elif campo_dian in ["79. Valor Fletes USD", "80. Valor Seguros USD", "81. Valor Otros Gastos USD"]: coincide = diferencia_absoluta < 1.0 or diferencia_porcentual < 1.0
                    else: coincide = diferencia_absoluta < 1.0 and diferencia_porcentual < 1.0
                    
                    if coincide:
                        fila_totales[nombre_campo_di] = f"✅ {total_di:.2f}"
                        fila_totales[nombre_campo_subpartida] = f"✅ {total_subpartida:.2f}"
                    else:
                        fila_totales[nombre_campo_di] = f"❌ {total_di:.2f}"
                        fila_totales[nombre_campo_subpartida] = f"❌ {total_subpartida:.2f}"
                        tiene_errores_totales = True
                else:
                    fila_totales[nombre_campo_di] = f"{total_di:.2f}"
                    fila_totales[nombre_campo_subpartida] = f"{total_subpartida:.2f}"
            except:
                fila_totales[nombre_campo_di] = f"{total_di:.2f}"
                fila_totales[nombre_campo_subpartida] = f"{total_subpartida:.2f}"
        
        fila_totales["Resultado verificación"] = "❌ TOTALES NO COINCIDEN" if tiene_errores_totales else "✅ TOTALES CONFORME"
        reporte_filas.append(fila_totales)

    def _agregar_totales_una_subpartida(self, reporte_filas, datos_dian, fila_subpartida):
        """Agrega fila de totales (Una Subpartida)"""
        datos_dian_validos = self.obtener_filas_validas_para_totales(datos_dian)
        fila_totales = {"4. Número DI": "VALORES ACUMULADOS"}
        tiene_errores_totales = False
        
        for campo in self.campos_consistencia: fila_totales[campo] = "N/A"
        for campo, (campo_dian, _) in self.campos_comparacion_individual.items():
            fila_totales[f"{campo_dian} DI"] = "N/A"
            fila_totales[f"{campo_dian} Subpartida"] = "N/A"

        valor_bultos_di_consolidado = 0
        if '74. Número de Bultos' in datos_dian.columns:
            s_bultos = pd.to_numeric(datos_dian['74. Número de Bultos'], errors='coerce')
            unicos = s_bultos.dropna().unique()
            if len(unicos) > 0: valor_bultos_di_consolidado = unicos[0]
            
        valor_bultos_excel = 0
        if fila_subpartida is not None:
             if isinstance(fila_subpartida, dict):
                 valor_bultos_excel = fila_subpartida.get('numero_bultos', 0)
             elif hasattr(fila_subpartida, 'numero_bultos'):
                 valor_bultos_excel = fila_subpartida['numero_bultos']

        for campo, (campo_dian, _) in self.campos_bultos.items():
            nombre_campo_di = f"{campo_dian} DI"
            nombre_campo_subpartida = f"{campo_dian} Subpartida"
            try:
                diff = abs(valor_bultos_di_consolidado - float(valor_bultos_excel))
                txt_di = f"{valor_bultos_di_consolidado:.0f}"
                txt_excel = f"{float(valor_bultos_excel):.0f}"
                
                if diff == 0:
                    fila_totales[nombre_campo_di] = f"✅ {txt_di}"
                    fila_totales[nombre_campo_subpartida] = f"✅ {txt_excel}"
                else:
                    fila_totales[nombre_campo_di] = f"❌ {txt_di}"
                    fila_totales[nombre_campo_subpartida] = f"❌ {txt_excel}"
                    tiene_errores_totales = True
            except:
                fila_totales[nombre_campo_di] = str(valor_bultos_di_consolidado)
                fila_totales[nombre_campo_subpartida] = str(valor_bultos_excel)

        for campo, (campo_dian, _) in self.campos_acumulables.items():
            nombre_campo_di = f"{campo_dian} DI"
            nombre_campo_subpartida = f"{campo_dian} Subpartida"
            if campo_dian in datos_dian_validos.columns and len(datos_dian_validos) > 0:
                total_dian = datos_dian_validos[campo_dian].sum()
                valor_subpartida = fila_subpartida.get(campo, 0)
                
                if campo == 'cantidad':
                    fila_totales[nombre_campo_di] = total_dian
                    fila_totales[nombre_campo_subpartida] = valor_subpartida
                else:
                    try:
                        if self.es_valor_valido(valor_subpartida) and total_dian != 0:
                            diferencia_absoluta = abs(float(total_dian) - float(valor_subpartida))
                            diferencia_porcentual = (diferencia_absoluta / float(valor_subpartida)) * 100
                            coincide = False
                            if campo_dian == "78. Valor FOB USD": coincide = diferencia_absoluta < 1.0 or diferencia_porcentual < 1.0
                            elif campo_dian in ["79. Valor Fletes USD", "80. Valor Seguros USD", "81. Valor Otros Gastos USD"]: coincide = diferencia_absoluta < 1.0 or diferencia_porcentual < 1.0
                            else: coincide = diferencia_absoluta < 0.1 and diferencia_porcentual < 0.1
                            
                            if coincide:
                                fila_totales[nombre_campo_di] = f"✅ {total_dian:.2f}"
                                fila_totales[nombre_campo_subpartida] = f"✅ {valor_subpartida:.2f}"
                            else:
                                fila_totales[nombre_campo_di] = f"❌ {total_dian:.2f}"
                                fila_totales[nombre_campo_subpartida] = f"❌ {valor_subpartida:.2f}"
                                tiene_errores_totales = True
                        else:
                            fila_totales[nombre_campo_di] = f"{total_dian:.2f}"
                            fila_totales[nombre_campo_subpartida] = f"{valor_subpartida:.2f}"
                    except:
                        fila_totales[nombre_campo_di] = f"{total_dian:.2f}"
                        fila_totales[nombre_campo_subpartida] = f"{valor_subpartida:.2f}"
            else:
                fila_totales[nombre_campo_di] = "N/A"
                fila_totales[nombre_campo_subpartida] = "N/A"
        
        fila_totales["Resultado verificación"] = "❌ TOTALES NO COINCIDEN" if tiene_errores_totales else "✅ TOTALES CONFORME"
        reporte_filas.append(fila_totales)

    def _ordenar_columnas_reporte_con_di(self, df_reporte):
        columnas_base = ['4. Número DI']
        columnas_consistencia = [c for c in self.campos_consistencia.keys() if c in df_reporte.columns]
        
        columnas_individuales = []
        for campo, (campo_dian, _) in self.campos_comparacion_individual.items():
            if f"{campo_dian} DI" in df_reporte.columns:
                columnas_individuales.extend([f"{campo_dian} DI", f"{campo_dian} Subpartida"])
        
        columnas_subpartida = []
        for campo, (campo_dian, _) in self.campos_subpartida_arancelaria.items():
            if f"{campo_dian} DI" in df_reporte.columns:
                columnas_subpartida.extend([f"{campo_dian} DI", f"{campo_dian} Subpartida"])
        
        columnas_bultos = []
        for campo, (campo_dian, _) in self.campos_bultos.items():
            if f"{campo_dian} DI" in df_reporte.columns:
                columnas_bultos.extend([f"{campo_dian} DI", f"{campo_dian} Subpartida"])
        
        columnas_acumulables = []
        for campo, (campo_dian, _) in self.campos_acumulables.items():
            if f"{campo_dian} DI" in df_reporte.columns:
                columnas_acumulables.extend([f"{campo_dian} DI", f"{campo_dian} Subpartida"])
        
        columnas_finales = ['Resultado verificación']
        todas = columnas_base + columnas_consistencia + columnas_individuales + columnas_subpartida + columnas_bultos + columnas_acumulables + columnas_finales
        return [c for c in todas if c in df_reporte.columns]

    def generar_reporte_comparacion(self, datos_dian, datos_subpartidas, output_path):
        df_reporte = self.generar_reporte_tabular(datos_dian, datos_subpartidas)
        if not df_reporte.empty:
            try:
                for col in df_reporte.columns:
                    if '74. Número de Bultos' in col:
                        mask_totales = df_reporte['4. Número DI'].str.contains('VALORES ACUMULADOS', na=False)
                        df_reporte.loc[~mask_totales, col] = pd.to_numeric(df_reporte.loc[~mask_totales, col], errors='coerce')
                
                df_reporte.to_excel(output_path, index=False)
                print(f"💾 Reporte de comparación guardado en: {output_path}")
                self._mostrar_resumen_estadistico(df_reporte)
            except Exception as e:
                print(f"❌ Error al guardar el reporte de comparación: {e}")
        return df_reporte
    
    def _mostrar_resumen_estadistico(self, df_reporte):
        di_individuales = df_reporte[~df_reporte['4. Número DI'].str.contains('VALORES ACUMULADOS', na=False)]
        print(f"\n📈 RESUMEN ESTADÍSTICO:")
        print(f"   • Total DI procesadas: {len(di_individuales)}")
        conformes = len(di_individuales[di_individuales['Resultado verificación'] == '✅ CONFORME'])
        print(f"   • DI conformes: {conformes}")
        print(f"   • DI con diferencias: {len(di_individuales) - conformes}")
        
        totales_multiples = df_reporte[df_reporte['4. Número DI'] == 'VALORES ACUMULADOS (MÚLTIPLES SUBPARTIDAS)']
        if not totales_multiples.empty:
            res_tot = totales_multiples.iloc[0]['Resultado verificación']
            print(f"   • Totales múltiples subpartidas: {res_tot}")

# =============================================================================
# CLASE 3: EXTRACCIÓN DE EXCEL (SUBPARTIDAS)
# =============================================================================

class ExtractorSubpartidas:
    def __init__(self):
        self.datos_estandarizados = pd.DataFrame()
    
    def buscar_archivo_subpartidas(self, carpeta_base):
        patrones = ["*subpartida*.xlsx", "*subpartida*.xls", "*resumen*.xlsx", "*resumen*.xls", "*.xlsx"]
        for patron in patrones:
            archivos = glob.glob(os.path.join(carpeta_base, patron))
            for archivo in archivos:
                nombre_archivo = os.path.basename(archivo).lower()
                if not any(palabra in nombre_archivo for palabra in ['resultado', 'validacion', 'comparacion', 'reporte']):
                    return archivo
        return None
    
    def detectar_hoja_correcta(self, archivo_excel):
        try:
            hojas_disponibles = pd.ExcelFile(archivo_excel).sheet_names
            palabras_clave = ['subpartida', 'resumen', 'datos', '847156', 'hoja1', 'sheet1']
            for hoja in hojas_disponibles:
                hoja_lower = hoja.lower()
                if any(palabra in hoja_lower for palabra in palabras_clave):
                    try:
                        df_prueba = pd.read_excel(archivo_excel, sheet_name=hoja, nrows=5)
                        if len([col for col in df_prueba.columns if any(palabra in str(col).lower() for palabra in ['subpartida', 'descripcion', 'peso', 'pais', 'valor'])]) >= 3:
                            return hoja
                    except: continue
            for hoja in hojas_disponibles:
                try:
                    df_prueba = pd.read_excel(archivo_excel, sheet_name=hoja, nrows=5)
                    if len(df_prueba.columns) >= 5 and not df_prueba.empty: return hoja
                except: continue
            return hojas_disponibles[0] if hojas_disponibles else None
        except: return None
    
    def extraer_y_estandarizar(self, carpeta_base) -> pd.DataFrame:
        try:
            archivo_excel = self.buscar_archivo_subpartidas(carpeta_base)
            if not archivo_excel: return pd.DataFrame()
            hoja_correcta = self.detectar_hoja_correcta(archivo_excel)
            if not hoja_correcta: return pd.DataFrame()
            df = pd.read_excel(archivo_excel, sheet_name=hoja_correcta, header=0)
            return self._estandarizar_y_filtrar_columnas(df)
        except: return pd.DataFrame()
    
    def _estandarizar_y_filtrar_columnas(self, df: pd.DataFrame) -> pd.DataFrame:
        mapeo_columnas = {
            'SUBPARTIDA': 'subpartida', 'DESCRIPCION': 'descripcion',
            'PESO BRUTO': 'peso_bruto', 'PESO NETO': 'peso_neto', 
            'NUMERO BULTOS': 'numero_bultos', 'PAIS ORIGEN': 'pais_origen',
            'PAIS COMPRA': 'pais_compra', 'PAIS PROCEDENCIA': 'pais_procedencia',
            'PAIS DESTINO': 'pais_destino', 'VALOR_FLETES': 'valor_fletes',
            'VALOR_SEGURO': 'valor_seguro', 'OTROS_GASTOS': 'otros_gastos',
            'BANDERA': 'bandera', 'UNIDAD': 'unidad',
            'VALOR FOB': 'valor_fob', 'CANTIDAD': 'cantidad'
        }
        columnas_a_mantener = [col for col in mapeo_columnas.keys() if col in df.columns]
        df_filtrado = df[columnas_a_mantener].copy()
        df_renombrado = df_filtrado.rename(columns=mapeo_columnas)
        df_renombrado['subpartida'] = df_renombrado['subpartida'].astype(str)
        df_renombrado['descripcion'] = df_renombrado['descripcion'].astype(str)
        campos_numericos = ['peso_bruto', 'peso_neto', 'numero_bultos', 'valor_fletes', 
                           'valor_seguro', 'otros_gastos', 'valor_fob', 'cantidad']
        for campo in campos_numericos:
            if campo in df_renombrado.columns:
                df_renombrado[campo] = pd.to_numeric(df_renombrado[campo], errors='coerce')
        return df_renombrado

# =============================================================================
# CLASE 4: VALIDACIÓN DECLARACIÓN IMPORTACIÓN
# =============================================================================

class ValidadorDeclaracionImportacionCompleto:
    def __init__(self):
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        self.corrector_nombres = CorrectorNombres()
        self.CAMPOS_DI = {
            "5.": "5. Número de Identificación Tributaria (NIT)",
            "11.": "11. Apellidos y Nombres / Razón Social Importador",
            "42.": "42. No. Manifiesto de Carga", "43.": "43. Fecha Manifiesto de Carga",
            "44.": "44. No. Documento de Transporte", "45.": "45. Fecha Documento de Transporte",
            "51.": "51. No. Factura Comercial", "52.": "52. Fecha Factura Comercial",
            "132.": "132. No. Aceptación Declaración", "133.": "133. Fecha Aceptación",
            "134.": "134. Levante No.", "135.": "135. Fecha Levante"
        }
        self.MAPEOS_VALIDACION = {
            "5. Número de Identificación Tributaria (NIT)": {"codigo_formulario": "PROVEEDOR", "descripcion_esperada": "INFORMACION_PROVEEDOR", "tipo": "documento", "cambia_por_declaracion": False},
            "11. Apellidos y Nombres / Razón Social Importador": {"codigo_formulario": "PROVEEDOR", "descripcion_esperada": "INFORMACION_PROVEEDOR", "tipo": "documento", "cambia_por_declaracion": False},
            "42. No. Manifiesto de Carga": {"codigo_formulario": 93, "descripcion_esperada": "FORMULARIO DE SALIDA ZONA FRANCA", "tipo": "documento", "cambia_por_declaracion": False},
            "43. Fecha Manifiesto de Carga": {"codigo_formulario": 93, "descripcion_esperada": "FORMULARIO DE SALIDA ZONA FRANCA", "tipo": "fecha", "cambia_por_declaracion": False},
            "44. No. Documento de Transporte": {"codigo_formulario": [17, 91], "descripcion_esperada": "DOCUMENTO OF TRANSPORTE", "tipo": "documento", "cambia_por_declaracion": False},
            "45. Fecha Documento de Transporte": {"codigo_formulario": [17, 91], "descripcion_esperada": "DOCUMENTO OF TRANSPORTE", "tipo": "fecha", "cambia_por_declaracion": False},
            "51. No. Factura Comercial": {"codigo_formulario": 6, "descripcion_esperada": "FACTURA COMERCIAL", "tipo": "documento", "cambia_por_declaracion": True},
            "52. Fecha Factura Comercial": {"codigo_formulario": 6, "descripcion_esperada": "FACTURA COMERCIAL", "tipo": "fecha", "cambia_por_declaracion": True},
            "132. No. Aceptación Declaración": {"codigo_formulario": 9, "descripcion_esperada": "DECLARACION DE IMPORTACION", "tipo": "documento", "cambia_por_declaracion": True},
            "133. Fecha Aceptación": {"codigo_formulario": 9, "descripcion_esperada": "DECLARACION DE IMPORTACION", "tipo": "fecha", "cambia_por_declaracion": True},
            "134. Levante No.": {"codigo_formulario": 47, "descripcion_esperada": "AUTORIZACION DE LEVANTE", "tipo": "documento", "cambia_por_declaracion": True},
            "135. Fecha Levante": {"codigo_formulario": 47, "descripcion_esperada": "AUTORIZACION DE LEVANTE", "tipo": "fecha", "cambia_por_declaracion": True}
        }
        self.patrones = {
            "5. Número de Identificación Tributaria (NIT)": [r"5\s*\.?\s*N[uú]mero\s*de\s*Identificaci[oó]n\s*Tributaria\s*\(NIT\).*?([0-9]{6,12})", r"5\.\s*Número de Identificación Tributaria \(NIT\)[\s\S]*?(\d{6,12})"],
            "11. Apellidos y Nombres / Razón Social Importador": [r"11\s*\.?\s*Apellidos\s*y\s*nombres\s*o\s*Raz[oó]n\s*Social\s*\n?\s*\d{6,12}\s*\d?\s*([A-ZÁÉÍÓÚÑ0-9\s\.\-&/]+?)(?=\s*13\s*\.)", r"11\.\s*Apellidos y nombres o Razón Social[\s\S]*?\n\s*(\d{6,12}\s*\d?\s*[A-ZÁÉÍÓÚÑ0-9\s\.\-&/]+)"],
            "42. No. Manifiesto de Carga": [r"42\s*\.?\s*Manifiesto\s*de\s*carga[\s\S]*?No\.?\s*([A-Z0-9]+)"],
            "43. Fecha Manifiesto de Carga": [r"43\s*\.?\s*Año\s*[-\s]*Mes\s*[-\s]*Día.*?(\d{4}\s*[-]\s*\d{2}\s*[-]\s*\d{2})"],
            "44. No. Documento de Transporte": [r"44\s*\.?\s*Documento\s*de\s*transporte[\s\S]*?No\.?\s*[A-Z0-9\-]{3,}[\s\S]*?No\.?\s*((?:(?=[A-Z0-9-]*[A-Z])[A-Z0-9]+(?:-[A-Z0-9]+)*)|(?:[A-Z]+\s*[0-9]+(?:-[A-Z]+)?)|(?:[A-Z0-9]{7,})|(?:[0-9]{6,11}))(?:\s|[0-9]{4}|$)"],
            "45. Fecha Documento de Transporte": [r"45\s*\.?\s*Año.*?Día[\s\S]*?[0-9]{4}\s*-\s*[0-9]{2}\s*-\s*[0-9]{2}[\s\S]*?([0-9]{4}\s*-\s*[0-9]{2}\s*-\s*[0-9]{2})"],
            "51. No. Factura Comercial": [r"51\s*\.?\s*No\.?\s*de\s*factura[\s\S]*?\n\s*(?!(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4})|(?:\d{4}[-/]\d{1,2}[-/]\d{1,2}))([A-Z0-9]+(?:/[A-Z0-9]+)*(?:-[A-Z0-9]+)*(?:\s+(?=[A-Z0-9]*[A-Z])[A-Z0-9]+)?)"],
            "52. Fecha Factura Comercial": [r"52\s*\.\s*?Año\s*-\s*Mes\s*-\s*Día.*?\n(?:.*?[^\d\w-])?(\d{4}\s*-\s*\d{2}\s*-\s*\d{2})"],
            "132. No. Aceptación Declaración": [r"132\s*\.?\s*No\.?\s*Aceptaci[oó]n\s*declaraci[oó]n[\s\S]*?(\d{12,18})"],
            "133. Fecha Aceptación": [r"133\s*\.?\s*Fec*h?a:?\s*(\d{4}\s*[\-\s]*\d{2}\s*[\-\s]*\d{2}|\d{8})\b"],
            "134. Levante No.": [r"134\s*\.?\s*Levante\s*No\.?[\s\S]{0,300}?(\d{12,})"],
            "135. Fecha Levante": [r"135\s*\.?\s*Fecha[\s\S]{0,400}?(\d{4}\s*-\s*\d{2}\s*-\s*\d{2})"]
        }
        self.nit_proveedor = None
        self.nombre_proveedor = None
        self.facturas_emparejadas = {}
        self._cache_nombres = {}

    def buscar_archivo_formulario(self, carpeta):
        print(f"🔍 Buscando formulario FMM...")
        patrones_formulario = ["*Rpt_Impresion_Formulario*", "*FORMULARIO*", "*FMM*", "*.xlsx"]
        for patron in patrones_formulario:
            archivos = glob.glob(os.path.join(carpeta, patron))
            for archivo in archivos:
                nombre_archivo = os.path.basename(archivo)
                if "Cruce" not in nombre_archivo and "validacion" not in nombre_archivo.lower():
                    print(f"📁 Formulario encontrado: {nombre_archivo}")
                    return archivo
        print("❌ No se encontró archivo del formulario FMM")
        return None

    def extraer_proveedor_formulario(self, archivo_excel):
        try:
            print(f"👤 Extrayendo información del proveedor...")
            wb = load_workbook(archivo_excel, data_only=True)
            sheet = wb.active
            proveedor_encontrado = False
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and 'Proveedor/Cliente:' in str(cell.value):
                        texto = str(cell.value)
                        print(f"📋 Información encontrada: {texto}")
                        texto_limpio = texto.replace('Proveedor/Cliente:', '').strip()
                        for i, char in enumerate(texto_limpio):
                            if char == ' ' and texto_limpio[:i].replace(' ', '').isdigit():
                                self.nit_proveedor = texto_limpio[:i].replace(' ', '').replace('-', '').replace('.', '')
                                self.nombre_proveedor = texto_limpio[i:].strip(' -')
                                break
                        if not self.nit_proveedor and ' - ' in texto_limpio:
                            partes = texto_limpio.split(' - ', 1)
                            self.nit_proveedor = partes[0].replace(' ', '').replace('-', '').replace('.', '')
                            self.nombre_proveedor = partes[1]
                        if self.nit_proveedor and self.nombre_proveedor:
                            print("✅ PROVEEDOR VÁLIDO:")
                            print(f"   🆔 NIT: {self.nit_proveedor}")
                            print(f"   📛 Nombre: {self.nombre_proveedor}")
                            proveedor_encontrado = True
                        wb.close()
                        return proveedor_encontrado
            wb.close()
            return proveedor_encontrado
        except Exception as e:
            print(f"❌ ERROR al extraer proveedor: {e}")
            return False

    def extraer_anexos_formulario_robusto(self, archivo_excel):
        try:
            print(f"📖 Extrayendo anexos del formulario...")
            wb = load_workbook(archivo_excel, data_only=True)
            sheet = wb.active
            inicio_anexos = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=row, column=col).value and 'DETALLE DE LOS ANEXOS' in str(sheet.cell(row=row, column=col).value):
                        inicio_anexos = row
                        break
                if inicio_anexos: break
            
            if inicio_anexos is None:
                wb.close(); return pd.DataFrame()
            
            encabezados = {}
            fila_encabezados = inicio_anexos + 1
            for col in range(1, sheet.max_column + 1):
                valor = sheet.cell(row=fila_encabezados, column=col).value
                if valor:
                    valor_str = str(valor).strip().upper()
                    if 'CÓDIGO' in valor_str or 'CODIGO' in valor_str: encabezados['codigo'] = col
                    elif 'DESCRIPCIÓN' in valor_str or 'DESCRIPCION' in valor_str: encabezados['descripcion'] = col
                    elif 'DOCUMENTO' in valor_str: encabezados['documento'] = col
                    elif 'FECHA' in valor_str: encabezados['fecha'] = col
            
            if not encabezados: encabezados = {'codigo': 1, 'descripcion': 5, 'documento': 19, 'fecha': 34}
            datos_anexos = []
            fila_actual = fila_encabezados + 1
            for i in range(200):
                try:
                    col_code = encabezados.get('codigo', 1)
                    codigo = sheet.cell(row=fila_actual, column=col_code).value
                    if codigo is None or codigo == '':
                        if all(sheet.cell(row=fila_actual + j, column=col_code).value in [None, ''] for j in range(3)): break
                        fila_actual += 1; continue
                    
                    try:
                        codigo_str = str(codigo).strip().split('.')[0]
                        if codigo_str not in ['6', '9', '17', '47', '93', '91']:
                            fila_actual += 1; continue
                    except: fila_actual += 1; continue
                    
                    datos_anexos.append({
                        'Codigo': int(float(codigo_str)),
                        'Descripcion': sheet.cell(row=fila_actual, column=encabezados.get('descripcion', 5)).value,
                        'Documento': sheet.cell(row=fila_actual, column=encabezados.get('documento', 19)).value,
                        'Fecha': self.normalizar_fecha_dd_mm_aaaa(sheet.cell(row=fila_actual, column=encabezados.get('fecha', 34)).value, es_fecha=True),
                        'Fila_Excel': fila_actual, 'Usado': False
                    })
                    fila_actual += 1
                except: fila_actual += 1; continue
            wb.close()
            df_resultado = pd.DataFrame(datos_anexos)
            
            # === CORRECCIÓN DE TIPOS PARA VALIDACIÓN ESTRICTA ===
            if not df_resultado.empty:
                # Convertir Documento a string para que coincida con el PDF
                df_resultado['Documento'] = df_resultado['Documento'].astype(str).str.strip()
                
                print(f"✅ {len(df_resultado)} anexos encontrados")
                resumen = df_resultado.groupby('Codigo').agg({'Descripcion': 'first', 'Documento': 'count'}).reset_index()
                print("📊 Resumen por código:")
                for _, row in resumen.iterrows():
                    print(f"   • Código {row['Codigo']}: {row['Documento']} - {row['Descripcion']}")
                
                di_rows = df_resultado[df_resultado['Codigo'] == 9]
                lev_rows = df_resultado[df_resultado['Codigo'] == 47]
                
                di_dupes = di_rows[di_rows.duplicated('Documento', keep=False)]['Documento'].unique()
                lev_dupes = lev_rows[lev_rows.duplicated('Documento', keep=False)]['Documento'].unique()
                
                count_di = len(di_rows)
                count_lev = len(lev_rows)
                
                has_integrity_issues = len(di_dupes) > 0 or len(lev_dupes) > 0 or count_di != count_lev
                
                if has_integrity_issues:
                    print("\n🔍 VALIDACIÓN DE INTEGRIDAD:")
                    if len(di_dupes) > 0:
                        print(f"   ❌ {len(di_dupes)} DI duplicadas: {', '.join(map(str, di_dupes))}")
                    if len(lev_dupes) > 0:
                        print(f"   ❌ {len(lev_dupes)} Levantes duplicados: {', '.join(map(str, lev_dupes))}")
                    if count_di != count_lev:
                        print(f"   ❌ Desbalance: {count_di} DI vs {count_lev} Levantes")
                else:
                    print(f"✅ Balance correcto: {count_di} DI = {count_lev} Levantes")

            return df_resultado
        except Exception as e:
            print(f"❌ Error al extraer anexos: {e}"); return pd.DataFrame()

    def extraer_todas_declaraciones_pdf(self, pdf_path):
        texto_completo = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text(x_tolerance=3, y_tolerance=3)
                    if texto: texto_completo += texto + "\n\n"
        except: return []
        
        matches = list(re.finditer(r"4\s*\.?\s*N[uú]mero\s*de\s*formulario[\s\S]*?(\d{12,18})", texto_completo, re.IGNORECASE))
        declaraciones = []
        for i, match in enumerate(matches):
            end_pos = matches[i+1].start() if i < len(matches) - 1 else len(texto_completo)
            declaraciones.append(self.extraer_datos_declaracion_individual(texto_completo[match.start():end_pos], match.group(1)))
        return declaraciones

    def extraer_datos_declaracion_individual(self, texto, numero_formulario):
        datos = {'Numero_Formulario_Declaracion': numero_formulario, 'Archivo_PDF': os.path.basename(texto.split('\n')[0]) if texto else 'Desconocido'}
        for campo in self.CAMPOS_DI.values():
            if campo in self.patrones:
                valor = self.extraer_campo_individual(texto, self.patrones[campo], campo)
                if any(p in campo for p in ['Fecha', 'Aceptación', 'Levante']): valor = self.normalizar_fecha_dd_mm_aaaa(valor, es_fecha=True)
                datos[campo] = valor
        return datos

    def extraer_campo_individual(self, texto, patrones, campo_nombre=""):
        for patron in patrones:
            try:
                match = re.search(patron, texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                if match:
                    if match.groups():
                        for group_val in match.groups():
                            if group_val and group_val.strip():
                                return group_val.strip()
                    else:
                        return match.group(0).strip()
            except: continue
        return "NO ENCONTRADO"

    def normalizar_fecha_dd_mm_aaaa(self, fecha_str, es_fecha=True):
        if not fecha_str or fecha_str == "NO ENCONTRADO" or str(fecha_str).strip() == "": return "NO ENCONTRADO"
        if not es_fecha: return str(fecha_str).strip()
        try:
            if isinstance(fecha_str, datetime): return fecha_str.strftime('%d-%m-%Y')
            fecha_limpia = str(fecha_str).strip().replace(' ', '')
            if len(fecha_limpia) > 10 and fecha_limpia.isdigit(): return fecha_limpia
            for patron, formato in [(r'^(\d{4})(\d{2})(\d{2})$', '%Y%m%d'), (r'(\d{4})-(\d{1,2})-(\d{1,2})', '%Y-%m-%d'), (r'(\d{4})/(\d{1,2})/(\d{1,2})', '%Y/%m/%d'), (r'(\d{1,2})-(\d{1,2})-(\d{4})', '%d-%m-%Y'), (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%d/%m/%Y')]:
                if re.match(patron, fecha_limpia): return datetime.strptime(fecha_limpia, formato).strftime('%d-%m-%Y')
            return fecha_limpia
        except: return str(fecha_str)

    def _normalizar_factura(self, factura_str):
        if not factura_str or factura_str == "NO ENCONTRADO": return ""
        return re.sub(r'[^\w\/\-]', '', str(factura_str).strip().upper().replace(' ', ''))

    def _emparejar_facturas_completo(self, facturas_declaraciones, facturas_formulario):
        emparejamientos = {}
        if len(facturas_formulario) == 1:
            for di in facturas_declaraciones: emparejamientos[di] = facturas_formulario[0]
        else:
            facturas_disp = facturas_formulario.copy()
            for di, fac in facturas_declaraciones.items():
                norm = self._normalizar_factura(fac)
                for f_form in facturas_disp:
                    if norm == self._normalizar_factura(f_form):
                        emparejamientos[di] = f_form; facturas_disp.remove(f_form); break
            for di, fac in facturas_declaraciones.items():
                if di not in emparejamientos:
                    part = self._normalizar_factura(fac).split('/')[0]
                    for f_form in facturas_disp:
                        if part == self._normalizar_factura(f_form).split('/')[0]:
                            emparejamientos[di] = f_form; facturas_disp.remove(f_form); break
            unmatched = [d for d in facturas_declaraciones if d not in emparejamientos]
            for i, di in enumerate(unmatched):
                if i < len(facturas_disp): emparejamientos[di] = facturas_disp[i]
        
        for di in facturas_declaraciones:
            if di not in emparejamientos: emparejamientos[di] = facturas_formulario[0] if facturas_formulario else "NO ENCONTRADO"
        return emparejamientos

    def _comparar_nombres_optimizado(self, nombre_pdf, nombre_excel):
        key = f"{nombre_pdf}_{nombre_excel}"
        if key not in self._cache_nombres:
            self._cache_nombres[key] = self.corrector_nombres.comparar_por_letras(nombre_pdf, nombre_excel)
        return self._cache_nombres[key]

    def validar_campos_por_declaracion(self, datos_declaracion, anexos_formulario):
        if anexos_formulario.empty and not (self.nit_proveedor and self.nombre_proveedor): return pd.DataFrame()
        resultados = []
        di_num = datos_declaracion.get('Numero_Formulario_Declaracion', 'NO ENCONTRADO')
        nom_pdf = datos_declaracion.get("11. Apellidos y Nombres / Razón Social Importador", "NO ENCONTRADO")
        
        for campo, config in self.MAPEOS_VALIDACION.items():
            res = {'Campos DI a Validar': campo, 'Datos Declaración': 'NO ENCONTRADO', 'Datos Formulario': 'NO ENCONTRADO', 'Numero DI': di_num, 'Coincidencias': '❌ NO COINCIDE'}
            try:
                val_dec = datos_declaracion.get(campo, "NO ENCONTRADO")
                if config["tipo"] == "fecha" and val_dec != "NO ENCONTRADO": val_dec = self.normalizar_fecha_dd_mm_aaaa(val_dec)
                
                if campo == "11. Apellidos y Nombres / Razón Social Importador":
                    res['Datos Declaración'] = self.corrector_nombres.corregir_nombre(nom_pdf, self.nombre_proveedor if self.nombre_proveedor else "")
                else: res['Datos Declaración'] = val_dec
                
                if config["codigo_formulario"] == "PROVEEDOR":
                    if campo == "5. Número de Identificación Tributaria (NIT)":
                        res['Datos Formulario'] = self.nit_proveedor or 'NO ENCONTRADO'
                        res['Coincidencias'] = '✅ COINCIDE' if str(val_dec).strip() == str(self.nit_proveedor).strip() else '❌ NO COINCIDE'
                    elif campo == "11. Apellidos y Nombres / Razón Social Importador":
                        res['Datos Formulario'] = self.nombre_proveedor or 'NO ENCONTRADO'
                        if self.nombre_proveedor and nom_pdf != "NO ENCONTRADO":
                            res['Coincidencias'] = '✅ COINCIDE' if self._comparar_nombres_optimizado(nom_pdf, self.nombre_proveedor) else '❌ NO COINCIDE'
                else:
                    codes = config["codigo_formulario"] if isinstance(config["codigo_formulario"], list) else [config["codigo_formulario"]]
                    
                    # --- FILTRO ESTRICTO (TODOS COMO STRING) ---
                    if 9 in codes:
                        anexos = anexos_formulario[
                            (anexos_formulario['Codigo'].isin(codes)) & 
                            (anexos_formulario['Documento'] == str(di_num)) # Forzar string
                        ]
                    elif 47 in codes:
                        levante_num = datos_declaracion.get("134. Levante No.", "NO ENCONTRADO")
                        anexos = anexos_formulario[
                            (anexos_formulario['Codigo'].isin(codes)) & 
                            (anexos_formulario['Documento'] == str(levante_num))
                        ]
                    elif 93 in codes: # Manifiesto
                        manif_num = datos_declaracion.get("42. No. Manifiesto de Carga", "NO ENCONTRADO")
                        anexos = anexos_formulario[
                            (anexos_formulario['Codigo'].isin(codes)) & 
                            (anexos_formulario['Documento'] == str(manif_num))
                        ]
                    elif 17 in codes or 91 in codes: # Transporte
                        transp_num = datos_declaracion.get("44. No. Documento de Transporte", "NO ENCONTRADO")
                        anexos = anexos_formulario[
                            (anexos_formulario['Codigo'].isin(codes)) & 
                            (anexos_formulario['Documento'] == str(transp_num))
                        ]
                    elif 6 in codes: # Factura
                        if campo == "52. Fecha Factura Comercial":
                            fact_num = self.facturas_emparejadas.get(di_num)
                            if fact_num:
                                fact_norm = self._normalizar_factura(fact_num)
                                indices_validos = []
                                for idx, row in anexos_formulario[anexos_formulario['Codigo'].isin(codes)].iterrows():
                                    if self._normalizar_factura(row['Documento']) == fact_norm:
                                        indices_validos.append(idx)
                                anexos = anexos_formulario.loc[indices_validos]
                            else:
                                anexos = pd.DataFrame() 
                        else:
                            anexos = anexos_formulario[anexos_formulario['Codigo'].isin(codes)]
                    else:
                        anexos = anexos_formulario[anexos_formulario['Codigo'].isin(codes)]
                    
                    if anexos.empty: res['Datos Formulario'] = 'NO ENCONTRADO'
                    else:
                        if not config["cambia_por_declaracion"]:
                            anexo = anexos.iloc[0]
                            val_form = anexo['Documento'] if config["tipo"] == "documento" else anexo.get('Fecha', 'NO ENCONTRADO')
                            res['Datos Formulario'] = val_form
                            res['Coincidencias'] = '✅ COINCIDE' if str(val_dec).strip() == str(val_form).strip() else '❌ NO COINCIDE'
                        else:
                            if campo == "51. No. Factura Comercial":
                                if di_num in self.facturas_emparejadas:
                                    val_form = self.facturas_emparejadas[di_num]
                                    res['Datos Formulario'] = val_form
                                    res['Coincidencias'] = '✅ COINCIDE' if self._normalizar_factura(val_dec) == self._normalizar_factura(val_form) else '❌ NO COINCIDE'
                            else:
                                match = False
                                for _, a in anexos.iterrows():
                                    v = a['Documento'] if config["tipo"] == "documento" else a.get('Fecha', 'NO ENCONTRADO')
                                    if str(val_dec).strip() == str(v).strip():
                                        res['Datos Formulario'] = v; res['Coincidencias'] = '✅ COINCIDE'; match = True; break
                                if not match:
                                    res['Datos Formulario'] = anexos.iloc[0]['Documento'] if config["tipo"] == "documento" else anexos.iloc[0].get('Fecha', 'NO ENCONTRADO')
            except Exception as e: res['Datos Formulario'] = f'ERROR: {str(e)}'
            resultados.append(res)
        return pd.DataFrame(resultados)

    def procesar_validacion_completa(self, carpeta_pdf, archivo_salida=None):
        form_file = self.buscar_archivo_formulario(carpeta_pdf)
        if not form_file: return None
        self.extraer_proveedor_formulario(form_file)
        anexos = self.extraer_anexos_formulario_robusto(form_file)
        if anexos.empty and not (self.nit_proveedor and self.nombre_proveedor): return None
        
        todas_decs = []
        for pdf in glob.glob(os.path.join(carpeta_pdf, "*.pdf")):
            print(f"\n📄 Procesando PDF: {os.path.basename(pdf)}")
            decs = self.extraer_todas_declaraciones_pdf(pdf)
            todas_decs.extend(decs)
            print(f"📋 {len(decs)} declaraciones encontradas")
            
        facts_form = anexos[anexos['Codigo'] == 6]['Documento'].tolist()
        facts_decs = {d.get('Numero_Formulario_Declaracion'): d.get('51. No. Factura Comercial', 'NO ENCONTRADO') for d in todas_decs if d.get('Numero_Formulario_Declaracion')}
        self.facturas_emparejadas = self._emparejar_facturas_completo(facts_decs, facts_form)
        
        if not archivo_salida: archivo_salida = os.path.join(carpeta_pdf, "Resultado Validacion Anexos FMM vs DIM.xlsx")
        
        all_results = []
        err_count = 0
        self._cache_nombres = {}
        print(f"🔍 Validando {len(todas_decs)} declaraciones...")
        
        for d in todas_decs:
            res = self.validar_campos_por_declaracion(d, anexos)
            if not res.empty:
                if len(res[res['Coincidencias'] == '❌ NO COINCIDE']) > 0: err_count += 1
                all_results.append(res)
        
        if all_results:
            df = pd.concat(all_results, ignore_index=True)
            try:
                with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Validacion_Detallada', index=False)
                print(f"\n{'='*50}\n📊 RESUMEN FINAL DE VALIDACIÓN\n{'='*50}")
                print(f"   • Total declaraciones procesadas: {len(todas_decs)}")
                print(f"   • Declaraciones con errores: {err_count}")
                print(f"   • Declaraciones correctas: {len(todas_decs)-err_count}")
                
                if err_count == 0:
                    print(f"🎯 TODAS LAS {len(todas_decs)} DECLARACIONES SON CORRECTAS ✅")
                else:
                    print(f"⚠️  {err_count} declaraciones requieren revisión")

                print(f"💾 Resultados guardados en: {archivo_salida}")
                print(f"{'='*50}")
                return df
            except PermissionError:
                print(f"❌ Error: Permiso denegado al guardar {archivo_salida}. Cierre el archivo si está abierto.")
            except Exception as e: print(f"❌ Error al guardar Excel: {e}")
        return None

# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def main():
    CARPETA_BASE = r"E:\Users\Lenovo\Desktop\PYTHON\DI\Junior Deposito 401\SLIND 401\SLIND 401\SLI 850232"
    EXCEL_OUTPUT_COMPARACION = os.path.join(CARPETA_BASE, "Resultado Validación Subpartida vs DIM.xlsx")
    EXCEL_OUTPUT_ANEXOS = os.path.join(CARPETA_BASE, "Resultado Validacion Anexos FMM vs DIM.xlsx")
    
    try:
        print("🚀 INICIANDO PROCESO COMPLETO DE EXTRACCIÓN Y COMPARACIÓN INTEGRADO")
        print(f"{'='*120}")
        print(f"📁 Carpeta base: {CARPETA_BASE}")
        
        if not os.path.exists(CARPETA_BASE): print("❌ Carpeta no existe"); return

        print(f"\n{'='*60}")
        print("📊 EJECUTANDO: Comparación DIM vs Subpartida")
        print(f"{'='*60}")
        
        print("\n📄 EXTRACCIÓN DE DATOS DE PDFs (DIAN)...")
        datos_dian = ExtractorDIANSimplificado().procesar_multiples_dis(CARPETA_BASE)
        
        print("\n📊 EXTRACCIÓN DE DATOS DE EXCEL (SUBPARTIDAS)...")
        datos_sub = ExtractorSubpartidas().extraer_y_estandarizar(CARPETA_BASE)
        
        if datos_dian is not None and not datos_dian.empty:
            print(f"✅ Datos DIAN extraídos: {len(datos_dian)} registros")
        else:
            print("❌ No se pudieron extraer datos DIAN")
            
        if not datos_sub.empty:
            print(f"✅ Datos Subpartidas extraídos: {len(datos_sub)} registros")
        else:
            print("❌ No se pudieron extraer datos de subpartidas")

        if datos_dian is not None and not datos_dian.empty and not datos_sub.empty:
            print("\n🔍 COMPARANDO DATOS EXTRAÍDOS...")
            reporte_comp = ComparadorDatos().generar_reporte_comparacion(datos_dian, datos_sub, EXCEL_OUTPUT_COMPARACION)
        else: 
            print("❌ Datos insuficientes para comparación")
            reporte_comp = None

        print(f"\n{'='*60}")
        print("📋 EJECUTANDO: Validación Anexos FMM vs DIM")
        print(f"{'='*60}")
        
        res_val = ValidadorDeclaracionImportacionCompleto().procesar_validacion_completa(CARPETA_BASE, EXCEL_OUTPUT_ANEXOS)
        
        print(f"\n{'='*120}")
        print("🎯 PROCESO COMPLETADO EXITOSAMENTE")
        print(f"{'='*120}")
        
        print(f"\n📁 ARCHIVOS GENERADOS:")
        
        if reporte_comp is not None and not reporte_comp.empty:
            conteo_real = len(reporte_comp[~reporte_comp['4. Número DI'].str.contains('VALORES ACUMULADOS', na=False)])
            print(f"   ✅ {EXCEL_OUTPUT_COMPARACION}")
            print(f"      • {conteo_real} DI procesadas")
            
        if res_val is not None:
             print(f"   ✅ {EXCEL_OUTPUT_ANEXOS}")
             print(f"      • Validación de anexos completada")

        print(f"\n📊 RESUMEN EJECUCIÓN:")
        print(f"   • Comparación DIM vs Subpartida: {'✅ COMPLETADO' if reporte_comp is not None else '❌ ERROR'}")
        print(f"   • Validación Anexos FMM: {'✅ COMPLETADO' if res_val is not None else '❌ ERROR'}")

            
    except Exception as e:
        print(f"❌ Error general: {e}")
        import traceback; traceback.print_exc()

if __name__ == "__main__":
    main()
